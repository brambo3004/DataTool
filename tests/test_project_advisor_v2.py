from __future__ import annotations

import io

import pandas as pd

from iasset_tool.nwegendocument_export import (
    build_nwegendocument_concept_rows,
    build_nwegendocument_concept_workbook_bytes,
)
from iasset_tool.project_advisor_v2 import (
    build_project_advisor_proposal_table,
    build_project_advisor_run_report,
    build_project_advisor_worklist,
    summarize_project_advisor,
)


def test_project_advisor_splitst_datakwaliteit_van_actielijst() -> None:
    """Een lokaal ontbrekend besteknummer mag geen werklijstregel worden."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "v1",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.6-04.6",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "aandacht",
                "datakwaliteit_signalen": "enkele objecten missen Besteknummer",
                "lokale_afwijkingen": "",
                "fysiek_lengte_m": 3003.0,
                "naam_begin": 1.6,
                "naam_eind": 4.6,
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    worklist = build_project_advisor_worklist(advisor_table)

    assert advisor_table.iloc[0]["adviesstatus"] == "ok"
    assert advisor_table.iloc[0]["datakwaliteitstatus"] == "aandacht"
    assert advisor_table.iloc[0]["grensstatus"] == "ok"
    assert advisor_table.iloc[0]["eindadvies"] == "Akkoord; datakwaliteit controleren"
    assert worklist.empty


def test_project_advisor_iasset_verschil_komt_in_werklijst() -> None:
    """Een voorstel dat afwijkt van iASSET blijft een echte databeheeractie."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "v2",
                "onderhoudsproject_voorgesteld": "N398-HRB-00.8-01.2",
                "vergelijking_iasset_status": "aandacht",
                "status_voorstel": "aandacht",
                "fysiek_lengte_m": 348.0,
                "naam_begin": 0.8,
                "naam_eind": 1.2,
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    worklist = build_project_advisor_worklist(advisor_table)

    assert advisor_table.iloc[0]["adviesstatus"] == "aandacht"
    assert advisor_table.iloc[0]["iassetvergelijking"] == "aandacht"
    assert advisor_table.iloc[0]["eindadvies"] == "Controleer verschil met iASSET"
    assert list(worklist["voorstel_id"]) == ["v2"]


def test_project_advisor_microvoorstel_wordt_apart_gemarkeerd() -> None:
    """Een nulmeter- of microvoorstel hoort niet als regulier project te voelen."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "v3",
                "onderhoudsproject_voorgesteld": "N398-HRB-06.3-06.3",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "controleer",
                "fysiek_lengte_m": 0.0,
                "naam_begin": 6.3,
                "naam_eind": 6.3,
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    worklist = build_project_advisor_worklist(advisor_table)

    assert advisor_table.iloc[0]["voorstelcategorie"] == "micro-/eindzonevoorstel"
    assert advisor_table.iloc[0]["eindadvies"] == "Niet als regulier onderhoudsproject gebruiken"
    assert list(worklist["voorstel_id"]) == ["v3"]


def test_project_advisor_terminale_grens_komt_in_werklijst() -> None:
    """De laatste projectgrens bij een afwijkend hm-interval blijft een actiepunt."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "v_ok",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.6-04.6",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "aandacht",
                "fysiek_lengte_m": 3003.0,
                "naam_begin": 1.6,
                "naam_eind": 4.6,
                "begin_hm_interval_afwijking_m": 28.86,
            },
            {
                "voorstel_id": "v_end",
                "onderhoudsproject_voorgesteld": "N398-HRB-04.9-06.3",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "controleer",
                "fysiek_lengte_m": 1357.0,
                "naam_begin": 4.9,
                "naam_eind": 6.3,
                "eind_hm_interval_afwijking_m": -26.96,
                "eind_grensdiagnose": "hectometerinterval 06.2-06.3 is fysiek 73.0 m in plaats van 100.0 m",
            },
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    worklist = build_project_advisor_worklist(advisor_table)

    assert advisor_table.loc[advisor_table["voorstel_id"] == "v_ok", "in_werklijst"].iloc[0] == False
    assert advisor_table.loc[advisor_table["voorstel_id"] == "v_end", "eindadvies"].iloc[0] == "Controleer eindgrens op kaart"
    assert list(worklist["voorstel_id"]) == ["v_end"]


def test_project_advisor_samenvatting_en_werklijst() -> None:
    """De samenvatting telt voorstellen, werklijstregels en hm-intervallen."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "ok",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.0-01.2",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "ok",
                "fysiek_lengte_m": 200.0,
                "naam_begin": 1.0,
                "naam_eind": 1.2,
            },
            {
                "voorstel_id": "controle",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.2-01.4",
                "vergelijking_iasset_status": "aandacht",
                "status_voorstel": "aandacht",
                "fysiek_lengte_m": 200.0,
                "naam_begin": 1.2,
                "naam_eind": 1.4,
            },
        ]
    )
    intervals = pd.DataFrame(
        [
            {"status": "ok"},
            {"status": "aandacht"},
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    summary = summarize_project_advisor(advisor_table, intervals)
    worklist = build_project_advisor_worklist(advisor_table)

    assert summary["voorstellen"] == 2
    assert summary["geen_directe_actie"] == 1
    assert summary["werklijstregels"] == 1
    assert summary["hm_intervallen_afwijkend"] == 1
    assert list(worklist["voorstel_id"]) == ["controle"]

def test_project_advisor_runrapport_geeft_automatisch_oordeel() -> None:
    """Het runrapport voorkomt dat de gebruiker specifieke regels handmatig moet zoeken."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "ok",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.6-04.6",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "aandacht",
                "datakwaliteit_signalen": "enkele objecten missen Besteknummer",
                "fysiek_lengte_m": 3003.0,
                "naam_begin": 1.6,
                "naam_eind": 4.6,
            },
            {
                "voorstel_id": "verschil",
                "onderhoudsproject_voorgesteld": "N398-HRB-00.8-01.2",
                "vergelijking_iasset_status": "aandacht",
                "status_voorstel": "aandacht",
                "fysiek_lengte_m": 348.0,
                "naam_begin": 0.8,
                "naam_eind": 1.2,
            },
            {
                "voorstel_id": "micro",
                "onderhoudsproject_voorgesteld": "N398-HRB-06.3-06.3",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "controleer",
                "fysiek_lengte_m": 0.0,
                "naam_begin": 6.3,
                "naam_eind": 6.3,
            },
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    summary = summarize_project_advisor(advisor_table)
    report = build_project_advisor_run_report(
        advisor_table,
        summary,
        selected_road="N398",
        uploaded_wegassen_name="wegassen_paspoort.geojson",
        app_version="v0.36.2",
    )

    oordeel = report.loc[report["onderdeel"] == "Run-oordeel"].iloc[0]
    assert oordeel["waarde"] == "Bruikbaar als databeheeradvies met werklijst"
    assert oordeel["urgentie"] == "actie"
    assert "werk eerst de werklijst af" in oordeel["vervolgstap"]

    werklijst = report.loc[report["onderdeel"] == "Werklijstregels"].iloc[0]
    assert werklijst["waarde"] == 2

    verschillen = report.loc[report["onderdeel"] == "Verschillen met iASSET"].iloc[0]
    assert "Vergelijk deze met het N-wegendocument" not in verschillen["vervolgstap"]
    assert "kaartbeeld" in verschillen["vervolgstap"]
    assert "actuele iASSET-data" in verschillen["vervolgstap"]
    assert "werkblad, niet als waarheid" in verschillen["vervolgstap"]


def test_project_advisor_runrapport_geen_voorstellen_is_stop() -> None:
    """Zonder voorstellen moet de run niet als werkbasis worden gepresenteerd."""
    advisor_table = build_project_advisor_proposal_table(pd.DataFrame())
    summary = summarize_project_advisor(advisor_table)
    report = build_project_advisor_run_report(advisor_table, summary, selected_road="N398")

    oordeel = report.loc[report["onderdeel"] == "Run-oordeel"].iloc[0]
    assert oordeel["waarde"] == "Geen bruikbaar projectadvies"
    assert oordeel["urgentie"] == "stop"



def test_nwegendocument_concept_export_vult_werkbladkolommen() -> None:
    """De concept-export vertaalt voorstellen naar de kolommen van het N-wegendocument."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "hrb",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.6-04.6",
                "project_type": "HRB",
                "project_family": "HRB",
                "bestaande_onderhoudsprojecten": "N398-HRB-01.6-04.6",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "ok",
                "fysiek_begin_m": 0.0,
                "fysiek_eind_m": 3000.0,
                "fysiek_begin_km": 1.6,
                "fysiek_eind_km": 4.6,
                "fysiek_lengte_m": 3000.0,
                "naam_begin": 1.6,
                "naam_eind": 4.6,
                "technisch_profiel": "Soort verharding_N=SMA, Soort deklaag specifiek=SMA-NL 8A, Jaar aanleg=1990, Jaar deklaag=2020, Jaar conservering=<leeg>, Jaar herstrating=<leeg>",
            },
            {
                "voorstel_id": "bb",
                "onderhoudsproject_voorgesteld": "N398-BBLR-04.6-04.8",
                "project_type": "BBLR",
                "project_family": "BB",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "ok",
                "fysiek_begin_m": 3000.0,
                "fysiek_eind_m": 3200.0,
                "fysiek_begin_km": 4.6,
                "fysiek_eind_km": 4.8,
                "fysiek_lengte_m": 200.0,
                "naam_begin": 4.6,
                "naam_eind": 4.8,
            },
        ]
    )
    objects = pd.DataFrame(
        [
            {
                "voorstel_id": "hrb",
                "naam": "rijstrook",
                "Besteknummer": "20-01-W",
                "Soort deklaag specifiek": "SMA-NL 8A",
                "Jaar aanleg": 1990,
                "Jaar deklaag": 2020,
            },
            {
                "voorstel_id": "hrb",
                "naam": "rijstrook-2",
                "Besteknummer": "20-01-W",
                "Soort deklaag specifiek": "SMA-NL 8A",
                "Jaar aanleg": 1990,
                "Jaar deklaag": 2020,
            },
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    concept_rows = build_nwegendocument_concept_rows(advisor_table, objects)

    hrb = concept_rows[concept_rows["voorstel_id"] == "hrb"].iloc[0]
    bb = concept_rows[concept_rows["voorstel_id"] == "bb"].iloc[0]

    assert hrb["tabblad"] == "HRB"
    assert hrb["onderhoudscomplex_nieuw"] == "N398-HRB-01.6-04.6"
    assert hrb["knip_begin"] == 1600
    assert hrb["besteknummer"] == "20-01-W"
    assert hrb["verhardingsoort"] == "SMA-NL 8A"
    assert str(hrb["jaar_deklaag"]) == "2020"

    # Busbanen horen in het concept bij het parallelweg-tabblad.
    assert bb["tabblad"] == "PW"



def test_nwegendocument_concept_export_gebruikt_hectometrering_in_meters() -> None:
    """Relatieve route-meters mogen niet in het N-wegendocument terechtkomen."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "n354",
                "onderhoudsproject_voorgesteld": "N354-HRB-25.8-26.3",
                "project_type": "HRB",
                "project_family": "HRB",
                # Relatieve route-meters op de wegas:
                "fysiek_begin_m": 0.0,
                "fysiek_eind_m": 524.0,
                # Administratieve hectometrering:
                "fysiek_begin_km": 25.8,
                "fysiek_eind_km": 26.3,
                "fysiek_lengte_m": 524.0,
                "naam_begin": 25.8,
                "naam_eind": 26.3,
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    concept_rows = build_nwegendocument_concept_rows(advisor_table)
    row = concept_rows.iloc[0]

    assert row["knip_begin"] == 25800
    assert row["knip_einde"] == 26300
    assert row["verharding_begin"] == 25800
    assert row["verharding_einde"] == 26300

def test_nwegendocument_concept_export_maakt_xlsx_bytes() -> None:
    """De app kan een downloadbare Excel-export maken zonder templatebestand."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "hrb",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.6-04.6",
                "project_type": "HRB",
                "project_family": "HRB",
                "vergelijking_iasset_status": "ok",
                "status_voorstel": "ok",
                "fysiek_begin_m": 0.0,
                "fysiek_eind_m": 3000.0,
                "fysiek_begin_km": 1.6,
                "fysiek_eind_km": 4.6,
                "fysiek_lengte_m": 3000.0,
                "naam_begin": 1.6,
                "naam_eind": 4.6,
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    xlsx_bytes = build_nwegendocument_concept_workbook_bytes(
        advisor_table,
        selected_road="N398",
        app_version="v0.36.4",
    )

    assert xlsx_bytes[:2] == b"PK"

    workbook = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    assert "Samenvatting" in workbook.sheet_names
    assert "N398 (HRB)" in workbook.sheet_names
    exported = pd.read_excel(workbook, sheet_name="N398 (HRB)", header=None)
    assert "onderhoudscomplex oud" in exported.iloc[2].astype(str).tolist()
    assert "N398-HRB-01.6-04.6" in exported.astype(str).to_string()


def test_nwegendocument_export_fp_heeft_compact_fietspadformat() -> None:
    """FP volgt het compacte N-wegendocumentformat zonder knipkolommen."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "fp",
                "onderhoudsproject_voorgesteld": "N354-FPR-25.8-27.5",
                "project_type": "FPR",
                "project_family": "FP",
                "bestaande_onderhoudsprojecten": "N354-FPR-25.7-28.0",
                "fysiek_begin_km": 25.8,
                "fysiek_eind_km": 27.5,
                "fysiek_lengte_m": 1700.0,
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    xlsx_bytes = build_nwegendocument_concept_workbook_bytes(advisor_table, selected_road="N354")

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    worksheet = workbook["N354 (FP)"]

    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    assert headers == [
        "onderhoudscomplex oud",
        "onderhoudscomplex nieuw",
        "objecten",
        "locatie",
        "besteknummer",
        "verhardingsoort",
        "conservering",
        "jaar aanleg",
        "jaar deklaag",
        "jaar conservering",
        "jaar herstrating",
        "bijzonderheden",
    ]
    assert worksheet.max_column == 12
    assert worksheet["A2"].value == "N354-FPR-25.7-28.0"
    assert worksheet["B2"].value == "N354-FPR-25.8-27.5"


def test_nwegendocument_export_pw_zet_oud_en_nieuw_in_kolom_b_en_c() -> None:
    """PW volgt de bestaande layout met filterkolom, oud in B en nieuw in C."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "pw",
                "onderhoudsproject_voorgesteld": "N354-PWL-10.9-12.7",
                "project_type": "PWL",
                "project_family": "PW",
                "bestaande_onderhoudsprojecten": "N354-PWL-10.8-20.5",
                "fysiek_begin_km": 10.9,
                "fysiek_eind_km": 12.7,
                "fysiek_lengte_m": 1800.0,
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    xlsx_bytes = build_nwegendocument_concept_workbook_bytes(advisor_table, selected_road="N354")

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    worksheet = workbook["N354 (PW)"]

    assert worksheet["A1"].value == "filters obv oude complexen"
    assert worksheet["B1"].value == "onderhoudscomplex oud"
    assert worksheet["C1"].value == "onderhoudscomplex nieuw"
    assert worksheet["B5"].value == "N354-PWL-10.8-20.5"
    assert worksheet["C5"].value == "N354-PWL-10.9-12.7"
    assert worksheet["D5"].value == 10900
    assert worksheet["E5"].value == 12700


def test_nwegendocument_export_objecten_kolom_vult_geen_generieke_objectlijst() -> None:
    """De zichtbare objecten-kolom is alleen voor bijzondere objecten."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "hrb",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.6-04.6",
                "project_type": "HRB",
                "project_family": "HRB",
                "fysiek_begin_km": 1.6,
                "fysiek_eind_km": 4.6,
                "fysiek_lengte_m": 3000.0,
            }
        ]
    )
    objects = pd.DataFrame(
        [
            {"voorstel_id": "hrb", "naam": "rijstrook-1", "subthema": "rijstrook"},
            {"voorstel_id": "hrb", "naam": "fietspad-2", "subthema": "fietspad"},
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    concept_rows = build_nwegendocument_concept_rows(advisor_table, objects)

    assert concept_rows.iloc[0]["objecten"] == ""


def test_nwegendocument_export_objecten_kolom_vult_bijzondere_objecten() -> None:
    """Bijzondere objecten blijven zichtbaar in de objecten-kolom."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "hrb",
                "onderhoudsproject_voorgesteld": "N354-HRB-10.9-11.5",
                "project_type": "HRB",
                "project_family": "HRB",
                "fysiek_begin_km": 10.9,
                "fysiek_eind_km": 11.5,
                "fysiek_lengte_m": 600.0,
            }
        ]
    )
    objects = pd.DataFrame(
        [
            {"voorstel_id": "hrb", "naam": "rijstrook-1", "subthema": "rijstrook"},
            {"voorstel_id": "hrb", "naam": "Rotonde Sneek", "subthema": "rotonde"},
            {"voorstel_id": "hrb", "naam": "brugdek", "Type onderdeel": "brug"},
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    concept_rows = build_nwegendocument_concept_rows(advisor_table, objects)

    assert concept_rows.iloc[0]["objecten"] == "Rotonde Sneek, brugdek"


def test_nwegendocument_export_schrijft_objecttoewijzing_apart() -> None:
    """Volledige objectcontext staat apart en niet in de zichtbare objecten-kolom."""
    proposals = pd.DataFrame(
        [
            {
                "voorstel_id": "hrb",
                "onderhoudsproject_voorgesteld": "N398-HRB-01.6-04.6",
                "project_type": "HRB",
                "project_family": "HRB",
                "fysiek_begin_km": 1.6,
                "fysiek_eind_km": 4.6,
                "fysiek_lengte_m": 3000.0,
            }
        ]
    )
    objects = pd.DataFrame(
        [
            {
                "voorstel_id": "hrb",
                "naam": "rijstrook-1",
                "subthema": "rijstrook",
                "WKT": "LINESTRING (0 0, 1 1)",
            }
        ]
    )

    advisor_table = build_project_advisor_proposal_table(proposals)
    xlsx_bytes = build_nwegendocument_concept_workbook_bytes(
        advisor_table,
        proposal_objects=objects,
        selected_road="N398",
        app_version="v0.36.6",
    )

    workbook = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    assert "Objecttoewijzing_data" in workbook.sheet_names

    object_sheet = pd.read_excel(workbook, sheet_name="Objecttoewijzing_data")
    assert "naam" in object_sheet.columns
    assert "WKT" not in object_sheet.columns
