import io

import pandas as pd
from openpyxl import load_workbook

from iasset_tool.project_calibration import (
    build_project_calibration_tables,
    build_project_calibration_workbook_bytes,
    project_calibration_export_filename,
)


def _advisor_table() -> pd.DataFrame:
    """Maak een kleine Project Adviseur-tabel met bekende kalibratiesignalen."""
    return pd.DataFrame(
        [
            {
                "voorstel_id": "kort",
                "onderhoudsproject_voorgesteld": "N354-HRB-11.0-11.1",
                "project_type": "HRB",
                "project_family": "HRB",
                "fysiek_begin_km": 10.96,
                "fysiek_eind_km": 11.075,
                "fysiek_lengte_m": 116.0,
                "voorstelcategorie": "regulier projectvoorstel",
                "iassetvergelijking": "aandacht",
                "in_werklijst": True,
                "knipreden_begin": "harde technische profielknip: Jaar aanleg, Jaar deklaag",
                "knipreden_eind": "harde technische profielknip: Soort verharding_N, Soort deklaag specifiek",
                "harde_knipsignalen": "harde technische profielknip: Jaar aanleg, Jaar deklaag",
            },
            {
                "voorstel_id": "dup1",
                "onderhoudsproject_voorgesteld": "N354-HRB-12.8-12.9",
                "project_type": "HRB",
                "project_family": "HRB",
                "fysiek_begin_km": 12.8,
                "fysiek_eind_km": 12.85,
                "fysiek_lengte_m": 50.0,
                "voorstelcategorie": "micro-/eindzonevoorstel",
                "iassetvergelijking": "ok",
                "in_werklijst": True,
                "knipreden_begin": "start spoor",
                "knipreden_eind": "einde spoor",
            },
            {
                "voorstel_id": "dup2",
                "onderhoudsproject_voorgesteld": "N354-HRB-12.8-12.9",
                "project_type": "HRB",
                "project_family": "HRB",
                "fysiek_begin_km": 12.86,
                "fysiek_eind_km": 12.9,
                "fysiek_lengte_m": 40.0,
                "voorstelcategorie": "micro-/eindzonevoorstel",
                "iassetvergelijking": "ok",
                "in_werklijst": True,
                "knipreden_begin": "start spoor",
                "knipreden_eind": "einde spoor",
            },
        ]
    )


def test_project_calibration_tables_vinden_kalibratiesignalen() -> None:
    """Het kalibratierapport wijst korte voorstellen, dubbele namen en mismatches aan."""
    objects = pd.DataFrame(
        [
            {"voorstel_id": "kort", "sys_id": "1", "naam": "fietspad-1", "subthema": "fietspad"},
            {"voorstel_id": "kort", "sys_id": "2", "naam": "rijstrook-1", "subthema": "rijstrook"},
        ]
    )
    intervals = pd.DataFrame(
        [
            {
                "axis_id": "WA-N354",
                "hm_interval": "20.4-20.5",
                "status": "aandacht",
                "afwijking_m": -25.0,
                "melding": "afwijkend interval",
            }
        ]
    )

    tables = build_project_calibration_tables(
        _advisor_table(),
        objects,
        intervals,
        selected_road="N354",
        app_version="v0.37.0",
    )

    assert set(tables) == {
        "Samenvatting",
        "Kandidaat_samenvoegen",
        "Korte_reguliere_voorstellen",
        "Dubbele_projectnamen",
        "Objectfamilie_mismatch",
        "Knipreden_analyse",
        "Hectometerinterval_context",
    }
    assert len(tables["Korte_reguliere_voorstellen"]) == 1
    assert tables["Korte_reguliere_voorstellen"].iloc[0]["voorstel_id"] == "kort"
    assert tables["Dubbele_projectnamen"]["onderhoudsproject_voorgesteld"].nunique() == 1
    assert len(tables["Objectfamilie_mismatch"]) == 1
    assert tables["Objectfamilie_mismatch"].iloc[0]["object_familie"] == "FP"
    assert "Jaar aanleg" in set(tables["Knipreden_analyse"]["knipveld"])
    assert len(tables["Hectometerinterval_context"]) == 1


def test_project_calibration_workbook_maakt_xlsx_met_tabs() -> None:
    """De app kan het kalibratierapport als Excelbestand downloaden."""
    xlsx_bytes = build_project_calibration_workbook_bytes(
        _advisor_table(),
        pd.DataFrame([{"voorstel_id": "kort", "sys_id": "1", "subthema": "fietspad"}]),
        pd.DataFrame([{"status": "aandacht", "hm_interval": "20.4-20.5"}]),
        selected_road="N354",
        app_version="v0.37.0",
    )

    assert xlsx_bytes[:2] == b"PK"

    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    assert "Samenvatting" in workbook.sheetnames
    assert "Kandidaat_samenvoegen" in workbook.sheetnames
    assert "Objectfamilie_mismatch" in workbook.sheetnames
    assert workbook["Samenvatting"]["A1"].value == "onderdeel"
    assert workbook["Kandidaat_samenvoegen"]["A1"].value == "prioriteit"


def test_project_calibration_filename() -> None:
    """De bestandsnaam volgt dezelfde veilige naamconventie als de andere exports."""
    assert project_calibration_export_filename("N354") == "Projectadvies_Kalibratie_N354.xlsx"
