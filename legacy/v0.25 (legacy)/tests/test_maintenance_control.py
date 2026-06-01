import io

import pandas as pd

from iasset_tool.maintenance_control import (
    action_work_queue_summary,
    build_control_point_object_details,
    build_maintenance_control,
    build_maintenance_control_workbook,
    build_data_quality_report,
    data_quality_summary,
    filter_action_work_queue,
    ensure_action_work_queue_columns,
    merge_action_work_queue_edits,
    merge_previous_action_follow_up,
    build_mutation_suggestions,
    normalize_project_name,
    read_action_lists_safely,
    read_maintenance_exports,
    summarize_action_projects,
    summarize_action_progress,
    summarize_maintenance_projects,
)


def test_normalize_project_name_keeps_content_but_fixes_spacing_and_dash():
    assert normalize_project_name(" N398 – HRB - 01.6 - 04.6 ") == "N398-HRB-01.6-04.6"


def test_read_maintenance_excel_detects_header_row_and_project_column():
    buffer = io.BytesIO()

    raw = pd.DataFrame(
        [
            ["", "Onderhoud, overzicht", ""],
            ["", "", ""],
            ["Object nr:", "Project", "Maatregel Omschrijving"],
            ["VV-1", "N398-HRB-01.6-04.6", "2L DGD A"],
        ]
    )

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        raw.to_excel(writer, index=False, header=False, sheet_name="Worksheet")

    result = read_maintenance_exports((("onderhoud.xlsx", buffer.getvalue()),))

    assert result.dataframe.shape[0] == 1
    assert result.dataframe.iloc[0]["Onderhoudsproject"] == "N398-HRB-01.6-04.6"
    assert result.dataframe.iloc[0]["maatregel"] == "2L DGD A"
    assert any("kopregel gevonden op rij 3" in warning for warning in result.warnings)



def test_read_maintenance_csv_detects_header_after_title_rows():
    csv_text = (
        "Onderhoud, overzicht\n"
        "\n"
        "Object nr:;Project;Maatregel Omschrijving\n"
        "VV-1;N398-HRB-01.0-02.0;DGD\n"
    )

    result = read_maintenance_exports((("onderhoud.csv", csv_text.encode("utf-8")),))

    assert result.dataframe.shape[0] == 1
    assert result.dataframe.iloc[0]["objectnummer"] == "VV-1"
    assert result.dataframe.iloc[0]["Onderhoudsproject"] == "N398-HRB-01.0-02.0"


def test_build_maintenance_control_finds_ok_missing_and_orphan_projects():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,1",
            },
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "bermverharding",
                "Onderhoudsproject": "N398-HRB-02.0-03.0",
                "Metrering": "2,1",
            },
        ]
    )

    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-1",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "maatregel": "2L DGD A",
            },
            {
                "objectnummer": "VV-999",
                "Onderhoudsproject": "N398-HRB-03.0-04.0",
                "maatregel": "LVOv",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")

    statuses = dict(zip(result.comparison["onderhoudsproject"], result.comparison["status"]))
    assert statuses["N398-HRB-01.0-02.0"] == "OK_VOLLEDIG"
    assert statuses["N398-HRB-02.0-03.0"] == "ONTBREEKT_IN_ONDERHOUD"
    assert statuses["N398-HRB-03.0-04.0"] == "GEEN_PASPOORTOBJECTEN"
    assert result.summary["projecten_ok"] == 1
    assert result.summary["ontbreekt_in_onderhoud"] == 1
    assert result.summary["geen_paspoortobjecten"] == 1


def test_passport_summary_counts_primary_secondary_and_exempt_objects():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
            },
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "bermverharding",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
            },
            {
                "nummer": "VV-3",
                "Wegnummer": "N398",
                "subthema": "perron",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
            },
        ]
    )

    result = build_maintenance_control(
        passport_df,
        pd.DataFrame([{"Onderhoudsproject": "N398-HRB-01.0-02.0"}]),
        selected_road="N398",
    )

    row = result.passport_projects.iloc[0]
    assert row["paspoort_primaire_objecten"] == 1
    assert row["paspoort_secundaire_objecten"] == 2
    assert row["paspoort_uitzondering_objecten"] == 1


def test_maintenance_summary_filters_by_selected_road_from_project_name():
    df = pd.DataFrame(
        [
            {"Onderhoudsproject": "N398-HRB-01.0-02.0", "objectnummer": "VV-1"},
            {"Onderhoudsproject": "N354-HRB-01.0-02.0", "objectnummer": "VV-2"},
        ]
    )

    summary = summarize_maintenance_projects(df, selected_road="N398")

    assert len(summary) == 1
    assert summary.iloc[0]["onderhoudsproject"] == "N398-HRB-01.0-02.0"


def test_build_maintenance_control_detects_object_difference_and_wrong_road_object():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-26704",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "Metrering": "0,8",
            },
            {
                "nummer": "VV-N398-26707",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "Metrering": "1,0",
            },
        ]
    )

    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N398-26704",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "maatregel": "OAB",
            },
            {
                "objectnummer": "VV-N398-26707",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "maatregel": "OAB",
            },
            {
                "objectnummer": "VV-N389-00000008",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "maatregel": "OAB",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")

    row = result.comparison.iloc[0]
    assert row["status"] == "OBJECT_WEGNUMMER_VERDACHT"
    assert row["alleen_in_onderhoud"] == 1
    assert row["onderhoud_object_wegnummer_verdacht"] == 1

    diff_types = set(result.object_differences["verschiltype"])
    assert "ALLEEN_IN_ONDERHOUD" in diff_types
    assert "OBJECT_WEGNUMMER_VERDACHT" in diff_types


def test_invalid_metrering_is_ignored_for_hm_range_but_reported():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-38213",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "Metrering": "4,9",
            },
            {
                "nummer": "VV-N398-38214",
                "Wegnummer": "N398",
                "subthema": "inrit en doorsteek",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "Metrering": "4,,9",
            },
        ]
    )

    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N398-38213",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N398-38214",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")

    summary_row = result.passport_projects.iloc[0]
    assert summary_row["paspoort_hm_min"] == 4.9
    assert summary_row["paspoort_hm_max"] == 4.9
    assert summary_row["paspoort_ongeldige_metrering_aantal"] == 1

    comparison_row = result.comparison.iloc[0]
    assert comparison_row["status"] == "HM_BEREIK_VERDACHT"

    diff_types = set(result.object_differences["verschiltype"])
    assert "ONGELDIGE_METRERING_PASPOORT" in diff_types



def test_action_list_translates_missing_project_to_work_instruction():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            }
        ]
    )

    result = build_maintenance_control(passport_df, pd.DataFrame(), selected_road="N354")

    assert result.summary["acties"] == 1
    action = result.action_list.iloc[0]
    assert action["status"] == "ONTBREEKT_IN_ONDERHOUD"
    assert action["controlecategorie"] == "Project ontbreekt in onderhoudsexport"
    assert "VV-N354-100" in action["betrokken_objecten"]
    assert "Zoek het onderhoudsproject exact op" in action["voorgestelde_actie"]


def test_action_list_explains_wrong_road_object_and_invalid_hm():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-26704",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "Metrering": "0,8",
            },
            {
                "nummer": "VV-N398-38214",
                "Wegnummer": "N398",
                "subthema": "inrit en doorsteek",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "Metrering": "4,,9",
            },
        ]
    )

    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N398-26704",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "maatregel": "OAB",
            },
            {
                "objectnummer": "VV-N389-00000008",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "maatregel": "OAB",
            },
            {
                "objectnummer": "VV-N398-38214",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")

    assert set(result.action_list["status"]) == {"OBJECT_WEGNUMMER_VERDACHT", "HM_BEREIK_VERDACHT"}

    wrong_road_action = result.action_list[result.action_list["status"] == "OBJECT_WEGNUMMER_VERDACHT"].iloc[0]
    assert "VV-N389-00000008" in wrong_road_action["betrokken_objecten"]
    assert "andere N-weg" in wrong_road_action["controlecategorie"]

    hm_action = result.action_list[result.action_list["status"] == "HM_BEREIK_VERDACHT"].iloc[0]
    assert "VV-N398-38214" in hm_action["betrokken_objecten"]
    assert "corrigeer de metrering" in hm_action["voorgestelde_actie"]


def test_action_list_contains_practical_category_and_follow_up_columns():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            }
        ]
    )

    result = build_maintenance_control(passport_df, pd.DataFrame(), selected_road="N354")
    action = result.action_list.iloc[0]

    assert action["praktische_categorie"] == "oude_of_ontbrekende_projectnaam_controleren"
    assert action["beoordeling_databeheerder"] == ""
    assert action["afhandelstatus"] == "nieuw"
    assert action["actiehouder"] == ""
    assert action["opmerking_afhandeling"] == ""


def test_action_list_uses_practical_category_for_wrong_road_and_object_difference():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-22551",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-31.6-31.7",
                "Metrering": "31,6",
            },
            {
                "nummer": "VV-N354-300",
                "Wegnummer": "N354",
                "subthema": "fietspad",
                "Onderhoudsproject": "N354-FPR-25.7-28.0",
                "Metrering": "25,8",
            },
        ]
    )

    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N354-22551",
                "Onderhoudsproject": "N354-HRB-31.6-31.7",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N354-22552",
                "Onderhoudsproject": "N354-HRB-31.6-31.7",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N354-300",
                "Onderhoudsproject": "N354-FPR-25.7-28.0",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N392-42740",
                "Onderhoudsproject": "N354-FPR-25.7-28.0",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")
    categories_by_status = dict(zip(result.action_list["status"], result.action_list["praktische_categorie"]))

    assert categories_by_status["OBJECT_WEGNUMMER_VERDACHT"] == "wegnummer_objectpaspoort_of_grensgeval_controleren"
    assert categories_by_status["OBJECTVERSCHIL"] == "objectset_of_projecttype_controleren"


def test_previous_action_list_follow_up_is_copied_to_matching_action():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            }
        ]
    )

    previous_action_list = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N354-HRB-11.5-12.8",
                "project_norm": "N354-HRB-11.5-12.8",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "controlecategorie": "Project ontbreekt in onderhoudsexport",
                "praktische_categorie": "oude_of_ontbrekende_projectnaam_controleren",
                "beoordeling_databeheerder": "Oude projectnaam staat nog bij objecten.",
                "afhandelstatus": "te corrigeren in paspoort",
                "actiehouder": "Bram",
                "opmerking_afhandeling": "Na correctie opnieuw exporteren.",
            }
        ]
    )

    result = build_maintenance_control(
        passport_df,
        pd.DataFrame(),
        selected_road="N354",
        previous_action_list=previous_action_list,
    )

    action = result.action_list.iloc[0]
    assert action["beoordeling_databeheerder"] == "Oude projectnaam staat nog bij objecten."
    assert action["afhandelstatus"] == "te corrigeren in paspoort"
    assert action["actiehouder"] == "Bram"
    assert action["opmerking_afhandeling"] == "Na correctie opnieuw exporteren."
    assert result.summary["acties_met_overgenomen_beoordeling"] == 1
    assert any("eerdere beoordeling" in warning for warning in result.warnings)


def test_previous_action_list_does_not_copy_to_changed_status():
    action_list = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N354-HRB-11.5-12.8",
                "project_norm": "N354-HRB-11.5-12.8",
                "status": "OBJECTVERSCHIL",
                "controlecategorie": "Objectsets verschillen",
                "praktische_categorie": "objectset_of_projecttype_controleren",
                "beoordeling_databeheerder": "",
                "afhandelstatus": "nieuw",
                "actiehouder": "",
                "opmerking_afhandeling": "",
            }
        ]
    )
    previous_action_list = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N354-HRB-11.5-12.8",
                "project_norm": "N354-HRB-11.5-12.8",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "controlecategorie": "Project ontbreekt in onderhoudsexport",
                "praktische_categorie": "oude_of_ontbrekende_projectnaam_controleren",
                "beoordeling_databeheerder": "Deze oude melding mag niet meekomen.",
                "afhandelstatus": "afgehandeld",
            }
        ]
    )

    merged, copied = merge_previous_action_follow_up(action_list, previous_action_list)

    assert copied == 0
    assert merged.iloc[0]["afhandelstatus"] == "nieuw"
    assert merged.iloc[0]["beoordeling_databeheerder"] == ""


def test_read_action_lists_safely_reads_semicolon_csv():
    csv_text = (
        "onderhoudsproject;status;controlecategorie;praktische_categorie;"
        "beoordeling_databeheerder;afhandelstatus;actiehouder;opmerking_afhandeling\n"
        "N354-HRB-11.5-12.8;ONTBREEKT_IN_ONDERHOUD;Project ontbreekt in onderhoudsexport;"
        "oude_of_ontbrekende_projectnaam_controleren;Oude projectnaam;in onderzoek;Bram;Controle loopt\n"
    )

    df, warnings = read_action_lists_safely((("Fase4_Actielijst_N354.csv", csv_text.encode("utf-8-sig")),))

    assert len(df) == 1
    assert df.iloc[0]["afhandelstatus"] == "in onderzoek"
    assert any("1 regel" in warning for warning in warnings)


def test_previous_action_list_accepts_human_column_names_and_v0162_matching():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            }
        ]
    )
    csv_text = (
        "onderhoudsproject;project_norm;status;controlecategorie;beoordeling databeheerder;afhandel status\n"
        "N354-HRB-11.5-12.8;N354-HRB-11.5-12.8;ONTBREEKT_IN_ONDERHOUD;"
        "Project ontbreekt in onderhoudsexport;Projectnaam is vervallen.;te corrigeren in paspoort\n"
    )
    previous_action_list, warnings = read_action_lists_safely((("oude_actielijst.csv", csv_text.encode("utf-8")),))

    assert "beoordeling_databeheerder" in previous_action_list.columns
    assert "afhandelstatus" in previous_action_list.columns
    assert any("mist opvolgkolommen" in warning for warning in warnings)

    result = build_maintenance_control(
        passport_df,
        pd.DataFrame(),
        selected_road="N354",
        previous_action_list=previous_action_list,
    )

    action = result.action_list.iloc[0]
    assert action["beoordeling_databeheerder"] == "Projectnaam is vervallen."
    assert action["afhandelstatus"] == "te corrigeren in paspoort"


def test_action_work_queue_summary_counts_follow_up_statuses():
    action_list = pd.DataFrame(
        [
            {"ernst": "waarschuwing", "afhandelstatus": "nieuw"},
            {"ernst": "waarschuwing", "afhandelstatus": "in onderzoek"},
            {"ernst": "aandachtspunt", "afhandelstatus": "te corrigeren in paspoort"},
            {"ernst": "waarschuwing", "afhandelstatus": "afgehandeld"},
        ]
    )

    summary = action_work_queue_summary(action_list)

    assert summary["controlepunten"] == 4
    assert summary["waarschuwingen"] == 3
    assert summary["aandachtspunten"] == 1
    assert summary["nieuw"] == 1
    assert summary["in_onderzoek"] == 1
    assert summary["te_corrigeren"] == 1
    assert summary["afgehandeld"] == 1
    assert summary["open"] == 3


def test_filter_action_work_queue_filters_by_category_status_and_search():
    action_list = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N354-HRB-11.5-12.8",
                "ernst": "waarschuwing",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "praktische_categorie": "oude_of_ontbrekende_projectnaam_controleren",
                "afhandelstatus": "nieuw",
                "actiehouder": "Bram",
                "betrokken_objecten": "VV-N354-100",
                "uitleg": "Project ontbreekt in onderhoudsexport",
            },
            {
                "onderhoudsproject": "N354-FPR-25.7-28.0",
                "ernst": "waarschuwing",
                "status": "OBJECT_WEGNUMMER_VERDACHT",
                "praktische_categorie": "wegnummer_objectpaspoort_of_grensgeval_controleren",
                "afhandelstatus": "in onderzoek",
                "actiehouder": "Willem",
                "betrokken_objecten": "VV-N392-42740",
                "uitleg": "Objectnummer lijkt bij andere N-weg te horen",
            },
        ]
    )

    result = filter_action_work_queue(
        action_list,
        praktische_categorie="wegnummer_objectpaspoort_of_grensgeval_controleren",
        afhandelstatus="in onderzoek",
        zoektekst="N392",
    )

    assert len(result) == 1
    assert result.iloc[0]["onderhoudsproject"] == "N354-FPR-25.7-28.0"


def test_merge_action_work_queue_edits_updates_only_follow_up_columns():
    base = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N354-HRB-11.5-12.8",
                "project_norm": "N354-HRB-11.5-12.8",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "controlecategorie": "Project ontbreekt in onderhoudsexport",
                "praktische_categorie": "oude_of_ontbrekende_projectnaam_controleren",
                "ernst": "waarschuwing",
                "afhandelstatus": "nieuw",
                "beoordeling_databeheerder": "",
                "actiehouder": "",
                "opmerking_afhandeling": "",
            },
            {
                "onderhoudsproject": "N354-FPR-25.7-28.0",
                "project_norm": "N354-FPR-25.7-28.0",
                "status": "OBJECT_WEGNUMMER_VERDACHT",
                "controlecategorie": "Objectnummer lijkt bij andere N-weg te horen",
                "praktische_categorie": "wegnummer_objectpaspoort_of_grensgeval_controleren",
                "ernst": "waarschuwing",
                "afhandelstatus": "nieuw",
                "beoordeling_databeheerder": "",
                "actiehouder": "",
                "opmerking_afhandeling": "",
            },
        ]
    )

    edited_subset = base.iloc[[1]].copy()
    edited_subset.loc[edited_subset.index[0], "afhandelstatus"] = "verklaarbare uitzondering"
    edited_subset.loc[edited_subset.index[0], "beoordeling_databeheerder"] = "Grensgeval bij wegnummerovergang"
    edited_subset.loc[edited_subset.index[0], "actiehouder"] = "Bram"

    result = merge_action_work_queue_edits(base, edited_subset)

    unchanged = result[result["onderhoudsproject"] == "N354-HRB-11.5-12.8"].iloc[0]
    changed = result[result["onderhoudsproject"] == "N354-FPR-25.7-28.0"].iloc[0]

    assert unchanged["afhandelstatus"] == "nieuw"
    assert changed["afhandelstatus"] == "verklaarbare uitzondering"
    assert changed["beoordeling_databeheerder"] == "Grensgeval bij wegnummerovergang"
    assert changed["actiehouder"] == "Bram"
    assert changed["status"] == "OBJECT_WEGNUMMER_VERDACHT"


def test_missing_project_gets_possible_maintenance_match_by_hm_overlap():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N354-200",
                "Onderhoudsproject": "N354-HRB-10.8-20.2",
                "maatregel": "DGD",
            }
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")

    missing_row = result.comparison[result.comparison["status"] == "ONTBREEKT_IN_ONDERHOUD"].iloc[0]
    assert "N354-HRB-10.8-20.2" in missing_row["mogelijke_onderhoudsmatch"]
    assert missing_row["onderhoudsmatch_type"] == "hm_overlap_zelfde_categorie"
    assert result.summary["acties_met_mogelijke_projectmatch"] == 1

    action = result.action_list[result.action_list["status"] == "ONTBREEKT_IN_ONDERHOUD"].iloc[0]
    assert "N354-HRB-10.8-20.2" in action["mogelijke_onderhoudsmatch"]
    assert "Mogelijke onderhoudsmatch" in action["uitleg"]
    assert "hint" in action["voorgestelde_actie"]


def test_missing_project_does_not_match_other_road_or_different_category():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N924-1",
                "Onderhoudsproject": "N924-HRB-11.5-12.8",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N354-2",
                "Onderhoudsproject": "N354-FPR-11.5-12.8",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")

    action = result.action_list[result.action_list["status"] == "ONTBREEKT_IN_ONDERHOUD"].iloc[0]
    assert action["mogelijke_onderhoudsmatch"] == ""
    assert action["onderhoudsmatch_score"] == 0



def test_missing_project_prefers_replacement_name_when_same_objects_are_under_new_project():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,6",
            },
            {
                "nummer": "VV-N354-101",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "12,1",
            },
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N354-100",
                "Onderhoudsproject": "N354-HRB-10.8-20.2",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N354-101",
                "Onderhoudsproject": "N354-HRB-10.8-20.2",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")

    action = result.action_list[result.action_list["status"] == "ONTBREEKT_IN_ONDERHOUD"].iloc[0]
    assert action["mogelijke_vervangende_projectnaam"] == "N354-HRB-10.8-20.2"
    assert action["onderhoudsmatch_type"] == "zelfde_objecten_mogelijke_nieuwe_projectnaam"
    assert action["vervangende_projectnaam_score"] >= 100
    assert "objectnummer" in action["vervangende_projectnaam_criteria"]
    assert "alleen een hint" in action["vervangende_projectnaam_uitleg"]

    suggestion = result.mutation_suggestions[
        result.mutation_suggestions["voorsteltype"] == "PROJECTNAAM_PASPOORT_CONTROLEREN"
    ].iloc[0]
    assert suggestion["mogelijke_vervangende_projectnaam"] == "N354-HRB-10.8-20.2"
    assert suggestion["matchscore"] >= 100
    assert "objectnummer" in suggestion["matchcriteria"]
    assert suggestion["automatisch_doorvoeren"] == "nee"


def test_missing_project_can_match_old_nonstandard_project_name_with_spaces():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354 HRB 11.5 12.8",
                "Metrering": "11,6",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N354-200",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")
    action = result.action_list[result.action_list["status"] == "ONTBREEKT_IN_ONDERHOUD"].iloc[0]

    assert action["mogelijke_vervangende_projectnaam"] == "N354-HRB-11.5-12.8"
    assert action["onderhoudsmatch_type"] == "hm_overlap_zelfde_categorie"
    assert "projectnaam" in action["vervangende_projectnaam_criteria"]


def test_missing_project_uses_passport_hm_and_subthema_when_old_name_has_no_range():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354 oud onderhoudscomplex west",
                "Metrering": "11,7",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N354-200",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")
    action = result.action_list[result.action_list["status"] == "ONTBREEKT_IN_ONDERHOUD"].iloc[0]

    assert action["mogelijke_vervangende_projectnaam"] == "N354-HRB-11.5-12.8"
    assert action["onderhoudsmatch_type"] == "hm_overlap_zelfde_categorie"
    assert "paspoortmetrering" in action["vervangende_projectnaam_criteria"]


def test_action_work_queue_keeps_possible_match_columns():
    action_list = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N354-HRB-11.5-12.8",
                "ernst": "waarschuwing",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "praktische_categorie": "oude_of_ontbrekende_projectnaam_controleren",
                "aantal_objecten": 1,
                "betrokken_objecten": "VV-N354-100",
                "mogelijke_onderhoudsmatch": "N354-HRB-10.8-20.2",
                "onderhoudsmatch_type": "hm_overlap_zelfde_categorie",
                "onderhoudsmatch_uitleg": "Mogelijke match op basis van hm-overlap.",
                "afhandelstatus": "nieuw",
            }
        ]
    )

    queue = ensure_action_work_queue_columns(action_list)

    assert "mogelijke_onderhoudsmatch" in queue.columns
    assert "onderhoudsmatch_type" in queue.columns
    assert queue.iloc[0]["mogelijke_onderhoudsmatch"] == "N354-HRB-10.8-20.2"



def test_mutation_suggestions_are_created_for_missing_project_with_match_hint():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N354-200",
                "Onderhoudsproject": "N354-HRB-10.8-20.2",
                "maatregel": "DGD",
            }
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")
    suggestions = result.mutation_suggestions

    project_suggestion = suggestions[
        suggestions["voorsteltype"] == "PROJECTNAAM_PASPOORT_CONTROLEREN"
    ].iloc[0]

    assert project_suggestion["huidige_waarde"] == "N354-HRB-11.5-12.8"
    assert project_suggestion["voorgestelde_waarde"] == "N354-HRB-10.8-20.2"
    assert project_suggestion["zekerheid"] == "match_hint"
    assert project_suggestion["voorstelstatus"] == "concept_voorstel"
    assert bool(project_suggestion["alleen_na_controle"]) is True
    assert bool(project_suggestion["menselijke_controle_verplicht"]) is True
    assert project_suggestion["automatisch_doorvoeren"] == "nee"
    assert "Niet automatisch doorvoeren" in project_suggestion["veiligheidsmelding"]
    assert result.summary["mutatievoorstellen"] == len(suggestions)


def test_mutation_suggestions_include_invalid_hm_and_wrong_road_object():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-38214",
                "Wegnummer": "N398",
                "subthema": "inrit en doorsteek",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "Metrering": "4,,9",
            },
            {
                "nummer": "VV-N398-26704",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "Metrering": "0,8",
            },
        ]
    )

    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N398-38214",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N398-26704",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "maatregel": "OAB",
            },
            {
                "objectnummer": "VV-N389-00000008",
                "Onderhoudsproject": "N398-HRB-00.8-01.1",
                "maatregel": "OAB",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")
    voorsteltypes = set(result.mutation_suggestions["voorsteltype"])

    assert "METRERING_PASPOORT_CORRIGEREN" in voorsteltypes
    assert "WEGNUMMER_OBJECT_CONTROLEREN" in voorsteltypes

    hm_row = result.mutation_suggestions[
        result.mutation_suggestions["voorsteltype"] == "METRERING_PASPOORT_CORRIGEREN"
    ].iloc[0]
    assert hm_row["objectnummer"] == "VV-N398-38214"
    assert hm_row["huidige_waarde"] == "4,,9"
    assert hm_row["zekerheid"] == "handmatig_bepalen"

    road_row = result.mutation_suggestions[
        result.mutation_suggestions["voorsteltype"] == "WEGNUMMER_OBJECT_CONTROLEREN"
    ].iloc[0]
    assert road_row["objectnummer"] == "VV-N389-00000008"
    assert road_row["huidige_waarde"] == "N389"
    assert road_row["voorgestelde_waarde"] == "N398"


def test_mutation_suggestions_always_include_hard_safety_columns():
    comparison = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N354-HRB-11.5-12.8",
                "project_norm": "N354-HRB-11.5-12.8",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "mogelijke_onderhoudsmatch": "N354-HRB-10.8-20.2",
            }
        ]
    )

    suggestions = build_mutation_suggestions(comparison, pd.DataFrame())

    assert not suggestions.empty
    assert set(
        [
            "voorstelstatus",
            "alleen_na_controle",
            "menselijke_controle_verplicht",
            "automatisch_doorvoeren",
            "veiligheidsmelding",
        ]
    ).issubset(suggestions.columns)
    assert suggestions["voorstelstatus"].eq("concept_voorstel").all()
    assert suggestions["alleen_na_controle"].eq(True).all()
    assert suggestions["menselijke_controle_verplicht"].eq(True).all()
    assert suggestions["automatisch_doorvoeren"].eq("nee").all()


def test_action_list_adds_duiding_for_likely_error_and_possible_old_project_name():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-38214",
                "Wegnummer": "N398",
                "subthema": "inrit en doorsteek",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "Metrering": "4,,9",
            },
            {
                "nummer": "VV-N354-100",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-11.5-12.8",
                "Metrering": "11,5",
            },
        ]
    )

    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N398-38214",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N354-200",
                "Onderhoudsproject": "N354-HRB-10.8-20.2",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df)

    hm_action = result.action_list[result.action_list["status"] == "HM_BEREIK_VERDACHT"].iloc[0]
    assert hm_action["duiding"] == "waarschijnlijke_fout_in_paspoort"

    missing_action = result.action_list[result.action_list["status"] == "ONTBREEKT_IN_ONDERHOUD"].iloc[0]
    assert missing_action["duiding"] == "mogelijke_oude_projectnaam"


def test_filter_action_work_queue_can_filter_by_duiding():
    action_list = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N398-HRB-04.8-04.9",
                "ernst": "aandachtspunt",
                "status": "HM_BEREIK_VERDACHT",
                "praktische_categorie": "metrering_paspoort_corrigeren",
                "duiding": "waarschijnlijke_fout_in_paspoort",
                "afhandelstatus": "nieuw",
            },
            {
                "onderhoudsproject": "N354-FPR-43.8-44.9",
                "ernst": "waarschuwing",
                "status": "OBJECT_WEGNUMMER_VERDACHT",
                "praktische_categorie": "wegnummer_objectpaspoort_of_grensgeval_controleren",
                "duiding": "fout_of_grensgeval_controleren",
                "afhandelstatus": "nieuw",
            },
        ]
    )

    result = filter_action_work_queue(action_list, duiding="fout_of_grensgeval_controleren")

    assert len(result) == 1
    assert result.iloc[0]["onderhoudsproject"] == "N354-FPR-43.8-44.9"


def test_mutation_suggestions_include_duiding_but_stay_safe():
    comparison = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N398-HRB-04.8-04.9",
                "project_norm": "N398-HRB-04.8-04.9",
                "status": "HM_BEREIK_VERDACHT",
            }
        ]
    )
    object_differences = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N398-HRB-04.8-04.9",
                "project_norm": "N398-HRB-04.8-04.9",
                "objectnummer": "VV-N398-38214",
                "verschiltype": "ONGELDIGE_METRERING_PASPOORT",
                "bron": "paspoortexport",
                "metrering": "4,,9",
            }
        ]
    )

    suggestions = build_mutation_suggestions(comparison, object_differences)

    assert "duiding" in suggestions.columns
    assert suggestions.iloc[0]["duiding"] == "waarschijnlijke_fout_in_paspoort"
    assert suggestions.iloc[0]["automatisch_doorvoeren"] == "nee"


def test_action_list_adds_human_duiding_group_and_explanation():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-38214",
                "Wegnummer": "N398",
                "subthema": "inrit en doorsteek",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "Metrering": "4,,9",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N398-38214",
                "Onderhoudsproject": "N398-HRB-04.8-04.9",
                "maatregel": "DGD",
            }
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")
    action = result.action_list.iloc[0]

    assert action["duiding"] == "waarschijnlijke_fout_in_paspoort"
    assert action["duiding_groep"] == "waarschijnlijke_fout"
    assert "objectpaspoort" in action["duiding_uitleg"]
    assert result.summary["acties_waarschijnlijke_fout"] == 1


def test_filter_action_work_queue_can_filter_by_duiding_group():
    action_list = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N398-HRB-04.8-04.9",
                "ernst": "aandachtspunt",
                "status": "HM_BEREIK_VERDACHT",
                "praktische_categorie": "metrering_paspoort_corrigeren",
                "duiding": "waarschijnlijke_fout_in_paspoort",
                "duiding_groep": "waarschijnlijke_fout",
                "afhandelstatus": "nieuw",
            },
            {
                "onderhoudsproject": "N354-HRB-31.6-31.7",
                "ernst": "waarschuwing",
                "status": "OBJECTVERSCHIL",
                "praktische_categorie": "objectset_of_projecttype_controleren",
                "duiding": "koppeling_of_exportmoment_controleren",
                "duiding_groep": "objectkoppeling_controleren",
                "afhandelstatus": "nieuw",
            },
        ]
    )

    result = filter_action_work_queue(action_list, duiding_groep="objectkoppeling_controleren")

    assert len(result) == 1
    assert result.iloc[0]["onderhoudsproject"] == "N354-HRB-31.6-31.7"


def test_mutation_suggestions_include_duiding_group_and_explanation():
    comparison = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N398-HRB-04.8-04.9",
                "project_norm": "N398-HRB-04.8-04.9",
                "status": "HM_BEREIK_VERDACHT",
            }
        ]
    )
    object_differences = pd.DataFrame(
        [
            {
                "onderhoudsproject": "N398-HRB-04.8-04.9",
                "project_norm": "N398-HRB-04.8-04.9",
                "objectnummer": "VV-N398-38214",
                "verschiltype": "ONGELDIGE_METRERING_PASPOORT",
                "bron": "paspoortexport",
                "metrering": "4,,9",
            }
        ]
    )

    suggestions = build_mutation_suggestions(comparison, object_differences)

    assert "duiding_groep" in suggestions.columns
    assert "duiding_uitleg" in suggestions.columns
    assert suggestions.iloc[0]["duiding_groep"] == "waarschijnlijke_fout"
    assert suggestions.iloc[0]["automatisch_doorvoeren"] == "nee"


def test_network_control_detects_wrong_object_road_from_project_name():
    """Areaalbrede controle gebruikt projectnaam als verwacht wegnummer."""
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-00001",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354-HRB-01.0-02.0",
                "Metrering": "1,1",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-N354-00001",
                "Onderhoudsproject": "N354-HRB-01.0-02.0",
                "maatregel": "DGD",
            },
            {
                "objectnummer": "VV-N392-99999",
                "Onderhoudsproject": "N354-HRB-01.0-02.0",
                "maatregel": "DGD",
            },
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road=None)

    assert result.summary["wegen_gecontroleerd"] == 1
    assert result.comparison.iloc[0]["wegnummer"] == "N354"
    assert result.comparison.iloc[0]["status"] == "OBJECT_WEGNUMMER_VERDACHT"
    assert "wegnummer" in result.action_list.columns
    assert result.action_list.iloc[0]["wegnummer"] == "N354"

    wrong_road = result.object_differences[
        result.object_differences["verschiltype"] == "OBJECT_WEGNUMMER_VERDACHT"
    ].iloc[0]
    assert wrong_road["object_wegnummer_vermoed"] == "N392"
    assert wrong_road["geselecteerde_weg"] == "N354"


def test_filter_action_work_queue_can_filter_by_wegnummer():
    action_list = pd.DataFrame(
        [
            {
                "wegnummer": "N354",
                "onderhoudsproject": "N354-HRB-01.0-02.0",
                "ernst": "waarschuwing",
                "status": "OBJECTVERSCHIL",
                "praktische_categorie": "objectset_of_projecttype_controleren",
                "afhandelstatus": "nieuw",
            },
            {
                "wegnummer": "N398",
                "onderhoudsproject": "N398-HRB-01.0-02.0",
                "ernst": "waarschuwing",
                "status": "OBJECTVERSCHIL",
                "praktische_categorie": "objectset_of_projecttype_controleren",
                "afhandelstatus": "nieuw",
            },
        ]
    )

    result = filter_action_work_queue(action_list, wegnummer="N398")

    assert len(result) == 1
    assert result.iloc[0]["onderhoudsproject"] == "N398-HRB-01.0-02.0"


def test_build_maintenance_control_workbook_contains_expected_sheets_and_follow_up_values():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-02.0-03.0",
                "Metrering": "2,0",
            },
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-1",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "maatregel": "DGD",
            }
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")
    edited_action_list = result.action_list.copy()
    edited_action_list.loc[:, "afhandelstatus"] = "in onderzoek"
    edited_action_list.loc[:, "actiehouder"] = "Databeheer Grijs"

    workbook_bytes = build_maintenance_control_workbook(
        result,
        action_list=edited_action_list,
        scope_label="N398",
    )

    assert workbook_bytes.startswith(b"PK")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert {
        "Samenvatting",
        "Werkvoorraad",
        "Mutatievoorstellen",
        "Resultaten",
        "Objectverschillen",
        "Paspoortprojecten",
        "Onderhoudsexport",
    }.issubset(set(wb.sheetnames))

    ws = wb["Werkvoorraad"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    afhandel_idx = header.index("afhandelstatus")
    actiehouder_idx = header.index("actiehouder")
    first_data_row = rows[1]
    assert first_data_row[afhandel_idx] == "in onderzoek"
    assert first_data_row[actiehouder_idx] == "Databeheer Grijs"


def test_build_maintenance_control_workbook_is_safe_when_no_actions_exist():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {
                "objectnummer": "VV-1",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "maatregel": "DGD",
            }
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")
    assert result.action_list.empty

    workbook_bytes = build_maintenance_control_workbook(result, scope_label="N398")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert "Samenvatting" in wb.sheetnames
    assert "Werkvoorraad" in wb.sheetnames


def test_action_progress_marks_existing_new_and_resolved_points():
    eerste_paspoort = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-02.0-03.0",
                "Metrering": "2,0",
            },
        ]
    )
    tweede_paspoort = eerste_paspoort[eerste_paspoort["nummer"] == "VV-2"].copy()
    onderhoud_leeg = pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"])

    eerste = build_maintenance_control(eerste_paspoort, onderhoud_leeg, selected_road="N398")
    vorige_actielijst = eerste.action_list.copy()
    vorige_actielijst.loc[
        vorige_actielijst["onderhoudsproject"] == "N398-HRB-02.0-03.0",
        "afhandelstatus",
    ] = "in onderzoek"
    vorige_actielijst.loc[
        vorige_actielijst["onderhoudsproject"] == "N398-HRB-02.0-03.0",
        "beoordeling_databeheerder",
    ] = "blijft open"

    tweede = build_maintenance_control(
        tweede_paspoort,
        onderhoud_leeg,
        selected_road="N398",
        previous_action_list=vorige_actielijst,
    )

    assert tweede.summary["controlepunten_bestaand"] == 1
    assert tweede.summary["controlepunten_nieuw"] == 0
    assert tweede.summary["controlepunten_opgelost"] == 1
    assert tweede.action_list.iloc[0]["voortgang_status"] == "bestaand_controlepunt"
    assert tweede.action_list.iloc[0]["afhandelstatus"] == "in onderzoek"
    assert tweede.action_list.iloc[0]["beoordeling_databeheerder"] == "blijft open"

    assert len(tweede.resolved_actions) == 1
    assert tweede.resolved_actions.iloc[0]["onderhoudsproject"] == "N398-HRB-01.0-02.0"
    assert tweede.resolved_actions.iloc[0]["voortgang_status"] == "opgelost_of_niet_meer_gevonden"


def test_action_progress_without_previous_list_marks_current_points_as_new():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
        ]
    )
    result = build_maintenance_control(
        passport_df,
        pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"]),
        selected_road="N398",
    )

    assert result.summary["controlepunten_nieuw"] == 1
    assert result.summary["controlepunten_bestaand"] == 0
    assert result.summary["controlepunten_opgelost"] == 0
    assert result.action_list.iloc[0]["voortgang_status"] == "nieuw_controlepunt"


def test_workbook_contains_resolved_action_sheet_when_previous_point_disappears():
    eerste_paspoort = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
        ]
    )
    tweede_paspoort = pd.DataFrame(
        [
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-02.0-03.0",
                "Metrering": "2,0",
            },
        ]
    )
    onderhoud_leeg = pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"])

    eerste = build_maintenance_control(eerste_paspoort, onderhoud_leeg, selected_road="N398")
    tweede = build_maintenance_control(
        tweede_paspoort,
        onderhoud_leeg,
        selected_road="N398",
        previous_action_list=eerste.action_list,
    )

    workbook_bytes = build_maintenance_control_workbook(tweede, scope_label="N398")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert "Opgelost" in wb.sheetnames
    rows = list(wb["Opgelost"].iter_rows(values_only=True))
    header = list(rows[0])
    assert "voortgang_status" in header
    assert rows[1][header.index("voortgang_status")] == "opgelost_of_niet_meer_gevonden"



def test_v024_progress_report_groups_new_existing_and_resolved_projectgroups():
    eerste_paspoort = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-02.0-03.0",
                "Metrering": "2,0",
            },
        ]
    )
    tweede_paspoort = pd.DataFrame(
        [
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-02.0-03.0",
                "Metrering": "2,0",
            },
            {
                "nummer": "VV-3",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-03.0-04.0",
                "Metrering": "3,0",
            },
        ]
    )
    onderhoud_leeg = pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"])

    eerste = build_maintenance_control(eerste_paspoort, onderhoud_leeg, selected_road="N398")
    tweede = build_maintenance_control(
        tweede_paspoort,
        onderhoud_leeg,
        selected_road="N398",
        previous_action_list=eerste.action_list,
    )

    project_report = tweede.progress_report[tweede.progress_report["scope"] == "onderhoudsproject"]
    conclusions = dict(zip(project_report["onderhoudsproject"], project_report["voortgang_conclusie"]))

    assert conclusions["N398-HRB-01.0-02.0"] == "opgelost_of_niet_meer_gevonden"
    assert conclusions["N398-HRB-02.0-03.0"] in {"blijft_open", "blijft_open_met_hoge_prioriteit"}
    assert conclusions["N398-HRB-03.0-04.0"] == "nieuw_in_deze_controle"
    assert tweede.summary["voortgang_nieuwe_projectgroepen"] == 1
    assert tweede.summary["voortgang_opgeloste_projectgroepen"] == 1


def test_v024_summarize_action_progress_can_be_used_directly():
    current = pd.DataFrame(
        [
            {
                "wegnummer": "N354",
                "onderhoudsproject": "N354-HRB-01.0-02.0",
                "project_norm": "N354-HRB-01.0-02.0",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "controlecategorie": "Project ontbreekt in onderhoudsexport",
                "praktische_categorie": "oude_projectnaam_of_migratie",
                "voortgang_status": "bestaand_controlepunt",
                "prioriteit": "hoog",
                "afhandelstatus": "in onderzoek",
                "actiehouder": "Databeheer Grijs",
            }
        ]
    )
    resolved = pd.DataFrame(
        [
            {
                "wegnummer": "N354",
                "onderhoudsproject": "N354-HRB-02.0-03.0",
                "project_norm": "N354-HRB-02.0-03.0",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "controlecategorie": "Project ontbreekt in onderhoudsexport",
                "praktische_categorie": "oude_projectnaam_of_migratie",
                "voortgang_status": "opgelost_of_niet_meer_gevonden",
            }
        ]
    )

    report = summarize_action_progress(current, resolved)
    assert {"weg", "onderhoudsproject"}.issubset(set(report["scope"]))
    project_rows = report[report["scope"] == "onderhoudsproject"].set_index("onderhoudsproject")
    assert project_rows.loc["N354-HRB-01.0-02.0", "voortgang_conclusie"] == "blijft_open_met_hoge_prioriteit"
    assert project_rows.loc["N354-HRB-01.0-02.0", "actiehouders"] == "Databeheer Grijs"
    assert project_rows.loc["N354-HRB-02.0-03.0", "voortgang_conclusie"] == "opgelost_of_niet_meer_gevonden"


def test_v024_workbook_contains_progress_report_sheet():
    eerste_paspoort = pd.DataFrame(
        [
            {
                "nummer": "VV-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
        ]
    )
    tweede_paspoort = pd.DataFrame(
        [
            {
                "nummer": "VV-2",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-02.0-03.0",
                "Metrering": "2,0",
            },
        ]
    )
    onderhoud_leeg = pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"])

    eerste = build_maintenance_control(eerste_paspoort, onderhoud_leeg, selected_road="N398")
    tweede = build_maintenance_control(
        tweede_paspoort,
        onderhoud_leeg,
        selected_road="N398",
        previous_action_list=eerste.action_list,
    )

    workbook_bytes = build_maintenance_control_workbook(tweede, scope_label="N398")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert "Voortgangsrapport" in wb.sheetnames
    rows = list(wb["Voortgangsrapport"].iter_rows(values_only=True))
    header = list(rows[0])
    assert "voortgang_conclusie" in header
    assert any(row[header.index("voortgang_conclusie")] == "opgelost_of_niet_meer_gevonden" for row in rows[1:])


def test_build_control_point_object_details_shows_passport_and_maintenance_sources():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-1",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Metrering": "1,0",
                "Situering": "Links",
                "Onderhoudsproject": "N354-HRB-01.0-02.0",
                "geometry": None,
            },
            {
                "nummer": "VV-N354-2",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Metrering": "1,1",
                "Situering": "Rechts",
                "Onderhoudsproject": "N354-HRB-01.0-02.0",
                "geometry": None,
            },
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {"objectnummer": "VV-N354-1", "Onderhoudsproject": "N354-HRB-01.0-02.0"},
            {"objectnummer": "VV-N354-3", "Onderhoudsproject": "N354-HRB-01.0-02.0"},
        ]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N354")
    action_row = result.action_list.iloc[0]

    details = build_control_point_object_details(
        passport_df,
        maintenance_df,
        action_row,
        result.object_differences,
        selected_road="N354",
    )

    assert set(details["objectnummer"]) == {"VV-N354-2", "VV-N354-3"}
    by_object = details.set_index("objectnummer")
    assert bool(by_object.loc["VV-N354-2", "in_paspoortexport"]) is True
    assert bool(by_object.loc["VV-N354-2", "in_onderhoudsexport"]) is False
    assert bool(by_object.loc["VV-N354-3", "in_paspoortexport"]) is False
    assert bool(by_object.loc["VV-N354-3", "in_onderhoudsexport"]) is True
    assert "ALLEEN_IN_PASPOORT" in by_object.loc["VV-N354-2", "verschiltype"]
    assert "ALLEEN_IN_ONDERHOUD" in by_object.loc["VV-N354-3", "verschiltype"]


def test_build_control_point_object_details_for_missing_project_lists_passport_objects():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Metrering": "1,0",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "geometry": None,
            }
        ]
    )
    maintenance_df = pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"])

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")
    action_row = result.action_list.iloc[0]

    details = build_control_point_object_details(
        passport_df,
        maintenance_df,
        action_row,
        result.object_differences,
        selected_road="N398",
    )

    assert list(details["objectnummer"]) == ["VV-N398-1"]
    assert bool(details.iloc[0]["in_paspoortexport"]) is True
    assert bool(details.iloc[0]["in_onderhoudsexport"]) is False

def test_v022_action_list_adds_priority_and_project_summary():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
            {
                "nummer": "VV-N398-2",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,1",
            },
        ]
    )
    maintenance_df = pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"])

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")

    row = result.action_list.iloc[0]
    assert row["prioriteit"] == "hoog"
    assert row["prioriteit_score"] >= 85
    assert "primaire paspoortobject" in row["prioriteit_uitleg"]
    assert "ontbreekt in de onderhoudsexport" in row["project_samenvatting"]
    assert result.summary["acties_prioriteit_hoog"] == 1


def test_v022_project_summary_groups_work_queue_and_recommends_next_step():
    action_list = pd.DataFrame(
        [
            {
                "wegnummer": "N398",
                "onderhoudsproject": "N398-HRB-01.0-02.0",
                "prioriteit": "hoog",
                "prioriteit_score": 95,
                "ernst": "waarschuwing",
                "status": "ONTBREEKT_IN_ONDERHOUD",
                "duiding_groep": "oude_projectnaam_of_migratie",
                "aantal_objecten": 12,
                "paspoort_primaire_objecten": 2,
                "paspoort_secundaire_objecten": 10,
                "mogelijke_vervangende_projectnaam": "N398-HRBR-01.0-02.0",
                "vervangende_projectnaam_score": 90,
                "afhandelstatus": "nieuw",
                "project_norm": "N398-HRB-01.0-02.0",
            }
        ]
    )

    summary = summarize_action_projects(action_list)

    assert list(summary["onderhoudsproject"]) == ["N398-HRB-01.0-02.0"]
    assert summary.iloc[0]["prioriteit"] == "hoog"
    assert summary.iloc[0]["open_controlepunten"] == 1
    assert summary.iloc[0]["mogelijke_vervangende_projectnaam"] == "N398-HRBR-01.0-02.0"
    assert "oude/nieuwe projectnaam" in summary.iloc[0]["korte_conclusie"]
    assert "Niets automatisch doorvoeren" in summary.iloc[0]["aanbevolen_volgende_stap"]


def test_v022_workbook_contains_project_summary_sheet():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N398-1",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            }
        ]
    )
    maintenance_df = pd.DataFrame(columns=["objectnummer", "Onderhoudsproject"])

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")
    workbook_bytes = build_maintenance_control_workbook(result, scope_label="N398")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert "Projectsamenvatting" in wb.sheetnames
    rows = list(wb["Projectsamenvatting"].iter_rows(values_only=True))
    header = list(rows[0])
    assert "prioriteit" in header
    assert rows[1][header.index("onderhoudsproject")] == "N398-HRB-01.0-02.0"




def test_v023_data_quality_report_flags_export_risks_before_control():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "VV-N354-1",
                "Wegnummer": "N354",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N354 oud complex",
                "Metrering": "3,,8",
            },
            {
                "nummer": "VV-N354-1",
                "Wegnummer": "N354",
                "subthema": "perron",
                "Onderhoudsproject": "N354-HRB-01.0-02.0",
                "Metrering": "1,0",
            },
        ]
    )
    maintenance_df = pd.DataFrame(
        [
            {"objectnummer": "", "Onderhoudsproject": "N354-HRB-01.0-02.0"},
            {"objectnummer": "VV-N354-1", "Onderhoudsproject": ""},
        ]
    )

    report = build_data_quality_report(passport_df, maintenance_df, selected_road="N354")

    codes = set(report["issue_code"])
    assert "PASPOORT_OBJECTNUMMER_DUBBEL" in codes
    assert "PASPOORT_METRERING_ONGELDIG" in codes
    assert "PASPOORT_UITZONDERINGSOBJECT_HEEFT_PROJECT" in codes
    assert "PASPOORT_PROJECTNAAM_FORMAT_AFWIJKEND" in codes
    assert "ONDERHOUD_PROJECTNAAM_LEEG" in codes
    assert "ONDERHOUD_OBJECTNUMMER_LEEG" in codes

    summary = data_quality_summary(report)
    assert summary["datakwaliteit_issues"] >= 6
    assert summary["datakwaliteit_waarschuwingen"] >= 2


def test_v023_maintenance_control_includes_data_quality_report_and_summary():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [{"objectnummer": "VV-N398-1", "Onderhoudsproject": "N398-HRB-01.0-02.0"}]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")

    assert not result.data_quality_report.empty
    assert "PASPOORT_OBJECTNUMMER_LEEG" in set(result.data_quality_report["issue_code"])
    assert result.summary["datakwaliteit_issues"] >= 1


def test_v023_workbook_contains_data_quality_sheet():
    passport_df = pd.DataFrame(
        [
            {
                "nummer": "",
                "Wegnummer": "N398",
                "subthema": "rijstrook",
                "Onderhoudsproject": "N398-HRB-01.0-02.0",
                "Metrering": "1,0",
            }
        ]
    )
    maintenance_df = pd.DataFrame(
        [{"objectnummer": "VV-N398-1", "Onderhoudsproject": "N398-HRB-01.0-02.0"}]
    )

    result = build_maintenance_control(passport_df, maintenance_df, selected_road="N398")
    workbook_bytes = build_maintenance_control_workbook(result, scope_label="N398")

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    assert "Datakwaliteit" in wb.sheetnames

    rows = list(wb["Datakwaliteit"].iter_rows(values_only=True))
    header = list(rows[0])
    assert "issue_code" in header
    assert any(row[header.index("issue_code")] == "PASPOORT_OBJECTNUMMER_LEEG" for row in rows[1:])
