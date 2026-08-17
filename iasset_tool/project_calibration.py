"""
Kalibratierapport voor Project Adviseur.

Deze module is bewust géén nieuwe projectvormingsengine. Zij maakt een
werkrapport waarmee de databeheerder en ontwikkelaar kunnen zien waar de
huidige onderhoudscomplexlogica waarschijnlijk te fijn of onlogisch knipt.

Waarom deze tussenlaag?
- N354 laat zien dat de tool end-to-end werkt, maar nog veel voorstellen maakt.
- We willen niet blind samenvoegregels aanpassen en daarmee N398 of echte
  onderhoudsverschillen stukmaken.
- Dit rapport wijst automatisch de belangrijkste kalibratieplekken aan:
  korte reguliere voorstellen, dubbele projectnamen, objectfamilie-mismatches,
  knipredenen en hectometerintervalcontext.
"""

from __future__ import annotations

import io
import re
from collections import Counter
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from .utils import clean_display_value, sanitize_filename


KALIBRATIE_SHEET_ORDER: tuple[str, ...] = (
    "Samenvatting",
    "Kandidaat_samenvoegen",
    "Korte_reguliere_voorstellen",
    "Dubbele_projectnamen",
    "Objectfamilie_mismatch",
    "Knipreden_analyse",
    "Hectometerinterval_context",
)

TECHNICAL_SIGNAL_FIELDS: tuple[str, ...] = (
    "Soort verharding_N",
    "Soort deklaag specifiek",
    "Jaar aanleg",
    "Jaar deklaag",
    "Jaar conservering",
    "Jaar herstrating",
    "Besteknummer",
)

SHORT_REGULAR_LIMIT_M = 200.0
VERY_SHORT_LIMIT_M = 100.0


def _text(value: Any) -> str:
    """Geef een veilige tekstwaarde zonder NaN-/None-ruis."""
    text = clean_display_value(value).strip()
    return "" if text.lower() in {"nan", "none", "nat", "<na>"} else text


def _number(value: Any) -> float | None:
    """Lees een getal robuust uit CSV-/DataFramewaarden."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    try:
        number = float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None
    return number if number == number else None


def _bool(value: Any) -> bool:
    """Lees een bool robuust uit echte bools en CSV-tekstwaarden."""
    if isinstance(value, bool):
        return value
    text = _text(value).lower()
    return text in {"true", "waar", "ja", "yes", "1"}


def _safe_sheet_name(name: str) -> str:
    """Maak een geldige Excel-tabbladnaam."""
    invalid = "[]:*?/\\"
    safe = "".join("_" if char in invalid else char for char in name)
    return safe[:31] or "Sheet"


def _proposal_family(row: pd.Series) -> str:
    """Normaliseer het voorgestelde projecttype naar een beheerfamilie."""
    project_type = _text(row.get("project_type", "")).upper()
    project_family = _text(row.get("project_family", "")).upper()
    proposed = _text(row.get("onderhoudsproject_voorgesteld", "")).upper()
    combined = " ".join(part for part in (project_type, project_family, proposed) if part)

    if project_type.startswith("HRB") or project_family == "HRB" or "-HRB" in proposed:
        return "HRB"
    if "FP" in combined:
        return "FP"
    if "BB" in combined or "BBLR" in combined:
        return "BB"
    if "LBP" in combined:
        return "LBP"
    if "PW" in combined:
        return "PW"
    return project_family or project_type or "ONBEKEND"


def _object_family_from_subtheme(row: pd.Series) -> str:
    """Bepaal de objectfamilie vanuit het oorspronkelijke object-subthema."""
    subthema = _text(row.get("subthema", row.get("Subthema", ""))).lower()

    if "rijstrook" in subthema:
        return "HRB"
    if "fietspad" in subthema:
        return "FP"
    if "parallelweg" in subthema:
        return "PW"
    if "busbaan" in subthema:
        return "BB"
    if "landbouwpad" in subthema:
        return "LBP"

    # Val terug op project_type/project_family in de objecttoewijzing als het
    # subthema ontbreekt. Dit blijft minder sterk bewijs dan het subthema.
    return _proposal_family(row)


def _accepted_object_families_for_proposal(proposal_family: str) -> set[str]:
    """
    Geef de verwachte primaire objectfamilies voor een voorstel.

    Busbaan en landbouwpad worden in het concept-N-wegendocument bij PW gezet,
    maar inhoudelijk blijven BB/LBP eigen primaire families. Daarom accepteren
    we ze niet automatisch in een gewoon PW-voorstel.
    """
    family = proposal_family.upper()
    if family in {"HRB", "FP", "PW", "BB", "LBP"}:
        return {family}
    return {family}


def _extract_signal_fields(*values: Any) -> list[str]:
    """Herken welke technische velden in knipredenen/signalen genoemd worden."""
    combined = " | ".join(_text(value) for value in values if _text(value)).lower()
    found: list[str] = []
    for field in TECHNICAL_SIGNAL_FIELDS:
        if field.lower() in combined:
            found.append(field)
    return found


def _split_signal_text(value: Any) -> list[str]:
    """Splits signaaltekst in compacte onderdelen."""
    text = _text(value)
    if not text:
        return []
    parts = re.split(r"\s*\|\s*|\s*;\s*", text)
    return [part.strip() for part in parts if part.strip()]


def _display_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Selecteer kolommen die bestaan en behoud een stabiele volgorde."""
    existing = [column for column in columns if column in df.columns]
    return df[existing].copy() if existing else df.copy()


def _build_short_regular_proposals(advisor_table: pd.DataFrame) -> pd.DataFrame:
    """Zoek korte voorstellen die nog als regulier voorstel zichtbaar zijn."""
    if advisor_table.empty:
        return pd.DataFrame()

    table = advisor_table.copy()
    table["_lengte_m_num"] = pd.to_numeric(table.get("fysiek_lengte_m"), errors="coerce")
    category = table.get("voorstelcategorie", pd.Series("", index=table.index)).fillna("").astype(str).str.lower()

    mask = (
        table["_lengte_m_num"].notna()
        & (table["_lengte_m_num"] <= SHORT_REGULAR_LIMIT_M)
        & (category == "regulier projectvoorstel")
    )
    result = table[mask].copy()
    if result.empty:
        return pd.DataFrame()

    result["kalibratie_signaal"] = result["_lengte_m_num"].map(
        lambda value: "zeer kort regulier voorstel" if value <= VERY_SHORT_LIMIT_M else "kort regulier voorstel"
    )
    result["advies_voor_v037x"] = (
        "Onderzoek of dit een echt zelfstandig onderhoudscomplex is of een kort technisch eilandje. "
        "Niet automatisch samenvoegen zonder kaart-/broncontrole."
    )
    result["technische_knipvelden"] = result.apply(
        lambda row: ", ".join(
            _extract_signal_fields(
                row.get("knipreden_begin", ""),
                row.get("knipreden_eind", ""),
                row.get("harde_knipsignalen", ""),
                row.get("zachte_signalen", ""),
            )
        ),
        axis=1,
    )

    columns = [
        "voorstel_id",
        "onderhoudsproject_voorgesteld",
        "project_type",
        "fysiek_begin_km",
        "fysiek_eind_km",
        "fysiek_lengte_m",
        "aantal_primaire_objecten",
        "bestaande_onderhoudsprojecten",
        "iassetvergelijking",
        "kalibratie_signaal",
        "technische_knipvelden",
        "knipreden_begin",
        "knipreden_eind",
        "harde_knipsignalen",
        "advies_voor_v037x",
    ]
    return _display_columns(result, columns)


def _build_duplicate_project_names(advisor_table: pd.DataFrame) -> pd.DataFrame:
    """Maak een tabel met projectnamen die meerdere voorstelregels krijgen."""
    if advisor_table.empty or "onderhoudsproject_voorgesteld" not in advisor_table.columns:
        return pd.DataFrame()

    table = advisor_table.copy()
    names = table["onderhoudsproject_voorgesteld"].map(_text)
    counts = names.value_counts()
    duplicate_names = set(counts[counts > 1].index)
    duplicate_names.discard("")

    if not duplicate_names:
        return pd.DataFrame()

    result = table[names.isin(duplicate_names)].copy()
    result["dubbele_projectnaam_aantal"] = result["onderhoudsproject_voorgesteld"].map(lambda value: int(counts.get(_text(value), 0)))
    result["advies_voor_v037x"] = (
        "Onderzoek of deze regels één zichtbaar onderhoudscomplex moeten worden of bewust parallel/deelcontext zijn."
    )

    columns = [
        "onderhoudsproject_voorgesteld",
        "dubbele_projectnaam_aantal",
        "voorstel_id",
        "project_type",
        "fysiek_begin_km",
        "fysiek_eind_km",
        "fysiek_lengte_m",
        "voorstelcategorie",
        "iassetvergelijking",
        "bestaande_onderhoudsprojecten",
        "knipreden_begin",
        "knipreden_eind",
        "advies_voor_v037x",
    ]
    return _display_columns(
        result.sort_values(["onderhoudsproject_voorgesteld", "fysiek_begin_km"], kind="stable"),
        columns,
    )


def _build_object_family_mismatch(advisor_table: pd.DataFrame, proposal_objects: pd.DataFrame) -> pd.DataFrame:
    """Zoek objecten die in een andere primaire familie zitten dan het voorstel."""
    if proposal_objects.empty or "voorstel_id" not in proposal_objects.columns:
        return pd.DataFrame()

    proposal_lookup = pd.DataFrame()
    if not advisor_table.empty and "voorstel_id" in advisor_table.columns:
        proposal_lookup = advisor_table[[
            column
            for column in [
                "voorstel_id",
                "onderhoudsproject_voorgesteld",
                "project_type",
                "project_family",
                "fysiek_begin_km",
                "fysiek_eind_km",
                "voorstelcategorie",
            ]
            if column in advisor_table.columns
        ]].copy()

    objects = proposal_objects.copy()
    if not proposal_lookup.empty:
        objects = objects.merge(proposal_lookup, on="voorstel_id", how="left", suffixes=("", "_voorstel"))

    # Projecttype/familie uit de voorstelrij heeft de voorkeur. Als die niet
    # bestaat, vallen we terug op de kolommen in de objecttoewijzing.
    objects["voorstel_familie"] = objects.apply(
        lambda row: _proposal_family(
            pd.Series(
                {
                    "project_type": row.get("project_type_voorstel", row.get("project_type", "")),
                    "project_family": row.get("project_family_voorstel", row.get("project_family", "")),
                    "onderhoudsproject_voorgesteld": row.get("onderhoudsproject_voorgesteld_voorstel", row.get("onderhoudsproject_voorgesteld", "")),
                }
            )
        ),
        axis=1,
    )
    objects["object_familie"] = objects.apply(_object_family_from_subtheme, axis=1)
    objects["verwachte_families"] = objects["voorstel_familie"].map(
        lambda family: ", ".join(sorted(_accepted_object_families_for_proposal(family)))
    )
    objects["familie_mismatch"] = objects.apply(
        lambda row: row["object_familie"] not in _accepted_object_families_for_proposal(row["voorstel_familie"]),
        axis=1,
    )

    result = objects[objects["familie_mismatch"]].copy()
    if result.empty:
        return pd.DataFrame()

    result["advies_voor_v037x"] = (
        "Onderzoek of dit object werkelijk bij dit voorstel hoort. "
        "Een afwijkende familie kan wijzen op kruispunt-/rotondecontext, maar ook op foutieve toewijzing."
    )

    columns = [
        "voorstel_id",
        "onderhoudsproject_voorgesteld",
        "project_type_voorstel",
        "project_family_voorstel",
        "voorstel_familie",
        "object_familie",
        "verwachte_families",
        "sys_id",
        "nummer",
        "naam",
        "subthema",
        "bestaand_onderhoudsproject",
        "fysiek_begin_km",
        "fysiek_eind_km",
        "toewijzing_status",
        "toewijzing_melding",
        "advies_voor_v037x",
    ]
    # Sommige merges leveren geen *_voorstel-kolommen op. Neem alternatieven mee.
    result = result.rename(
        columns={
            "project_type": "project_type_object",
            "project_family": "project_family_object",
        }
    )
    columns = [column for column in columns if column in result.columns]
    return result[columns].copy()


def _build_merge_candidates(
    short_regular: pd.DataFrame,
    duplicate_names: pd.DataFrame,
    object_mismatch: pd.DataFrame,
    advisor_table: pd.DataFrame,
) -> pd.DataFrame:
    """
    Combineer kalibratiesignalen tot één actielijst voor v0.37.x.

    Dit is géén automatische samenvoegactie. Het is een geprioriteerde lijst
    van voorstellen of objecten waarvoor de samenvoeg-/kniplogica waarschijnlijk
    aandacht vraagt.
    """
    rows: list[dict[str, Any]] = []

    for _, row in short_regular.iterrows():
        rows.append(
            {
                "prioriteit": 1 if _number(row.get("fysiek_lengte_m")) and _number(row.get("fysiek_lengte_m")) <= VERY_SHORT_LIMIT_M else 2,
                "type_kalibratiepunt": _text(row.get("kalibratie_signaal", "kort regulier voorstel")),
                "voorstel_id": _text(row.get("voorstel_id", "")),
                "onderhoudsproject_voorgesteld": _text(row.get("onderhoudsproject_voorgesteld", "")),
                "project_type": _text(row.get("project_type", "")),
                "fysiek_begin_km": row.get("fysiek_begin_km", ""),
                "fysiek_eind_km": row.get("fysiek_eind_km", ""),
                "fysiek_lengte_m": row.get("fysiek_lengte_m", ""),
                "reden": "Regulier voorstel is kort; mogelijk technisch eilandje of lokale overgang.",
                "advies_voor_v037x": _text(row.get("advies_voor_v037x", "")),
            }
        )

    if not duplicate_names.empty:
        for _, row in duplicate_names.drop_duplicates("onderhoudsproject_voorgesteld").iterrows():
            rows.append(
                {
                    "prioriteit": 1,
                    "type_kalibratiepunt": "dubbele projectnaam",
                    "voorstel_id": "",
                    "onderhoudsproject_voorgesteld": _text(row.get("onderhoudsproject_voorgesteld", "")),
                    "project_type": _text(row.get("project_type", "")),
                    "fysiek_begin_km": row.get("fysiek_begin_km", ""),
                    "fysiek_eind_km": row.get("fysiek_eind_km", ""),
                    "fysiek_lengte_m": row.get("fysiek_lengte_m", ""),
                    "reden": f"Projectnaam komt {_text(row.get('dubbele_projectnaam_aantal', 'meerdere'))} keer voor.",
                    "advies_voor_v037x": _text(row.get("advies_voor_v037x", "")),
                }
            )

    if not object_mismatch.empty:
        mismatch_counts = object_mismatch.groupby("voorstel_id", dropna=False).size().reset_index(name="aantal_mismatches")
        for _, count_row in mismatch_counts.iterrows():
            proposal_id = _text(count_row.get("voorstel_id", ""))
            proposal_row = pd.Series(dtype="object")
            if not advisor_table.empty and "voorstel_id" in advisor_table.columns:
                matches = advisor_table[advisor_table["voorstel_id"].map(_text) == proposal_id]
                if not matches.empty:
                    proposal_row = matches.iloc[0]
            rows.append(
                {
                    "prioriteit": 2,
                    "type_kalibratiepunt": "objectfamilie-mismatch",
                    "voorstel_id": proposal_id,
                    "onderhoudsproject_voorgesteld": _text(proposal_row.get("onderhoudsproject_voorgesteld", "")),
                    "project_type": _text(proposal_row.get("project_type", "")),
                    "fysiek_begin_km": proposal_row.get("fysiek_begin_km", ""),
                    "fysiek_eind_km": proposal_row.get("fysiek_eind_km", ""),
                    "fysiek_lengte_m": proposal_row.get("fysiek_lengte_m", ""),
                    "reden": f"{int(count_row.get('aantal_mismatches', 0))} object(en) hebben een afwijkende objectfamilie.",
                    "advies_voor_v037x": "Onderzoek of projectfamilie-toewijzing strenger moet worden of dat kruispuntcontext ontbreekt.",
                }
            )

    result = pd.DataFrame(rows)
    if result.empty:
        return pd.DataFrame(
            columns=[
                "prioriteit",
                "type_kalibratiepunt",
                "voorstel_id",
                "onderhoudsproject_voorgesteld",
                "project_type",
                "fysiek_begin_km",
                "fysiek_eind_km",
                "fysiek_lengte_m",
                "reden",
                "advies_voor_v037x",
            ]
        )

    return result.sort_values(["prioriteit", "project_type", "fysiek_begin_km"], kind="stable").reset_index(drop=True)


def _build_knip_reason_analysis(advisor_table: pd.DataFrame) -> pd.DataFrame:
    """Tel welke technische velden het vaakst als knipreden terugkomen."""
    if advisor_table.empty:
        return pd.DataFrame()

    rows: list[dict[str, Any]] = []
    for _, proposal in advisor_table.iterrows():
        proposal_id = _text(proposal.get("voorstel_id", ""))
        project_type = _text(proposal.get("project_type", "onbekend")) or "onbekend"
        proposed = _text(proposal.get("onderhoudsproject_voorgesteld", ""))

        signal_sources = {
            "knipreden_begin": proposal.get("knipreden_begin", ""),
            "knipreden_eind": proposal.get("knipreden_eind", ""),
            "harde_knipsignalen": proposal.get("harde_knipsignalen", ""),
            "zachte_signalen": proposal.get("zachte_signalen", ""),
            "bestek_signalen": proposal.get("bestek_signalen", ""),
            "datakwaliteit_signalen": proposal.get("datakwaliteit_signalen", ""),
        }
        fields = _extract_signal_fields(*signal_sources.values())
        if not fields:
            fields = ["geen technisch veld herkend"]

        for field in fields:
            rows.append(
                {
                    "knipveld": field,
                    "project_type": project_type,
                    "voorstel_id": proposal_id,
                    "onderhoudsproject_voorgesteld": proposed,
                    "fysiek_lengte_m": proposal.get("fysiek_lengte_m", ""),
                    "knipreden_begin": proposal.get("knipreden_begin", ""),
                    "knipreden_eind": proposal.get("knipreden_eind", ""),
                    "harde_knipsignalen": proposal.get("harde_knipsignalen", ""),
                }
            )

    detail = pd.DataFrame(rows)
    if detail.empty:
        return pd.DataFrame()

    summary = detail.groupby(["knipveld", "project_type"], dropna=False).agg(
        aantal_voorstellen=("voorstel_id", "nunique"),
        voorbeeld_project=("onderhoudsproject_voorgesteld", "first"),
    ).reset_index()
    total_proposals = max(int(advisor_table["voorstel_id"].nunique()) if "voorstel_id" in advisor_table.columns else len(advisor_table), 1)
    summary["percentage_van_voorstellen"] = (summary["aantal_voorstellen"] / total_proposals * 100).round(1)
    summary["interpretatie"] = summary["knipveld"].map(
        lambda field: (
            "Technisch profielveld: kandidaat voor inhoudelijke weging in samenvoeglogica."
            if field in TECHNICAL_SIGNAL_FIELDS
            else "Geen specifiek technisch veld herkend; bekijk detailvelden."
        )
    )
    return summary.sort_values(["aantal_voorstellen", "knipveld"], ascending=[False, True], kind="stable").reset_index(drop=True)


def _build_hectometer_interval_context(hectometer_intervals: pd.DataFrame) -> pd.DataFrame:
    """Selecteer alleen afwijkende hectometerintervallen als contextlaag."""
    if hectometer_intervals.empty or "status" not in hectometer_intervals.columns:
        return pd.DataFrame()

    table = hectometer_intervals.copy()
    status = table["status"].fillna("ok").astype(str).str.lower()
    result = table[status != "ok"].copy()
    if result.empty:
        return pd.DataFrame()

    columns = [
        "axis_id",
        "axis_naam",
        "Wegnummer",
        "hm_interval",
        "hm_van",
        "hm_tot",
        "verwachte_lengte_m",
        "gemeten_lengte_m",
        "afwijking_m",
        "afwijking_pct",
        "interval_factor",
        "status",
        "melding",
    ]
    return _display_columns(result, columns)


def _build_summary_sheet(
    advisor_table: pd.DataFrame,
    short_regular: pd.DataFrame,
    duplicate_names: pd.DataFrame,
    object_mismatch: pd.DataFrame,
    knip_analysis: pd.DataFrame,
    interval_context: pd.DataFrame,
    selected_road: str,
    app_version: str,
) -> pd.DataFrame:
    """Maak de samenvatting van het kalibratierapport."""
    proposals = len(advisor_table)
    worklist = int(advisor_table.get("in_werklijst", pd.Series(False, index=advisor_table.index)).map(_bool).sum()) if not advisor_table.empty else 0
    micro = int((advisor_table.get("voorstelcategorie", pd.Series("", index=advisor_table.index)).fillna("").astype(str).str.lower() == "micro-/eindzonevoorstel").sum()) if not advisor_table.empty else 0
    outside = int((advisor_table.get("voorstelcategorie", pd.Series("", index=advisor_table.index)).fillna("").astype(str).str.lower() == "buiten ijkbereik").sum()) if not advisor_table.empty else 0
    iasset_diff = int(advisor_table.get("iassetvergelijking", pd.Series("", index=advisor_table.index)).fillna("").astype(str).str.lower().isin(["aandacht", "controleer"]).sum()) if not advisor_table.empty else 0
    duplicate_project_count = int(duplicate_names["onderhoudsproject_voorgesteld"].nunique()) if not duplicate_names.empty and "onderhoudsproject_voorgesteld" in duplicate_names.columns else 0

    top_knip = ""
    if not knip_analysis.empty:
        top_row = knip_analysis.iloc[0]
        top_knip = f"{top_row.get('knipveld', '')} ({top_row.get('aantal_voorstellen', '')} voorstellen)"

    rows = [
        {
            "onderdeel": "Doel v0.37.0",
            "waarde": "Kalibratierapport, geen nieuwe projectlogica",
            "betekenis": "Dit rapport wijst automatisch aan waar de onderhoudscomplexlogica waarschijnlijk te fijn knipt.",
            "vervolgstap": "Gebruik deze output om v0.37.1 gericht te kiezen; pas nog niet blind samenvoegregels toe.",
        },
        {
            "onderdeel": "Weg",
            "waarde": selected_road or "onbekend",
            "betekenis": f"App-versie: {app_version or 'onbekend'}",
            "vervolgstap": "Bewaar dit rapport bij de bijbehorende Project Adviseur-run.",
        },
        {
            "onderdeel": "Projectvoorstellen",
            "waarde": proposals,
            "betekenis": "Aantal voorstellen in de huidige Project Adviseur-uitvoer.",
            "vervolgstap": "Bij hoge aantallen eerst naar korte voorstellen, dubbele namen en knipredenanalyse kijken.",
        },
        {
            "onderdeel": "Werklijstregels",
            "waarde": worklist,
            "betekenis": "Aantal voorstellen dat nu als actiepunt zichtbaar is.",
            "vervolgstap": "Niet automatisch reduceren; gebruik kalibratiesheets om de oorzaak te vinden.",
        },
        {
            "onderdeel": "iASSET-verschillen",
            "waarde": iasset_diff,
            "betekenis": "Verschillen met bestaande iASSET-indeling; iASSET blijft bron van waarheid, oude handmatige werkbladen niet.",
            "vervolgstap": "Beoordeel verschillen met kaartbeeld, actuele iASSET-data en broninformatie.",
        },
        {
            "onderdeel": "Korte reguliere voorstellen",
            "waarde": len(short_regular),
            "betekenis": f"Reguliere voorstellen van maximaal {SHORT_REGULAR_LIMIT_M:.0f} meter.",
            "vervolgstap": "Belangrijkste kandidaat voor samenvoeglogica: korte technische eilandjes herkennen.",
        },
        {
            "onderdeel": "Micro-/eindzonevoorstellen",
            "waarde": micro,
            "betekenis": "Voorstellen die al niet als regulier onderhoudsproject worden gepresenteerd.",
            "vervolgstap": "Controleer of deze categorie voldoende buiten de reguliere export blijft.",
        },
        {
            "onderdeel": "Buiten ijkbereik",
            "waarde": outside,
            "betekenis": "Voorstellen waarvan begin/eind buiten betrouwbaar ijkbereik ligt.",
            "vervolgstap": "Niet met onderhoudslogica oplossen; eerst kaart-/referentieascontext beoordelen.",
        },
        {
            "onderdeel": "Dubbele projectnamen",
            "waarde": duplicate_project_count,
            "betekenis": "Aantal voorgestelde projectnamen dat meerdere regels oplevert.",
            "vervolgstap": "Onderzoek of deze regels moeten worden samengevoegd of bewust parallel blijven.",
        },
        {
            "onderdeel": "Objectfamilie-mismatches",
            "waarde": len(object_mismatch),
            "betekenis": "Objecten waarvan subthemafamilie niet past bij voorsteltype.",
            "vervolgstap": "Kandidaat voor strengere projectfamilie-toewijzing of kruispuntcontext.",
        },
        {
            "onderdeel": "Belangrijkste knipsignaal",
            "waarde": top_knip or "niet bepaald",
            "betekenis": "Het technische veld dat het vaakst terugkomt in knipredenen.",
            "vervolgstap": "Gebruik dit om te bepalen welke profielwisselingen straks minder hard of contextgevoeliger moeten worden.",
        },
        {
            "onderdeel": "Afwijkende hm-intervallen",
            "waarde": len(interval_context),
            "betekenis": "Referentieas-/ijkcontext; niet automatisch onderhoudsprojectfout.",
            "vervolgstap": "Alleen gebruiken om grensbetrouwbaarheid uit te leggen.",
        },
    ]
    return pd.DataFrame(rows)


def build_project_calibration_tables(
    advisor_table: pd.DataFrame | None,
    proposal_objects: pd.DataFrame | None = None,
    hectometer_intervals: pd.DataFrame | None = None,
    *,
    selected_road: str = "",
    app_version: str = "",
) -> dict[str, pd.DataFrame]:
    """
    Bouw alle tabellen voor het kalibratierapport.

    De functie accepteert lege of incomplete DataFrames en crasht dan niet.
    Dat is belangrijk bij experimentele iASSET-exports met ontbrekende kolommen.
    """
    advisor = advisor_table.copy() if isinstance(advisor_table, pd.DataFrame) else pd.DataFrame()
    objects = proposal_objects.copy() if isinstance(proposal_objects, pd.DataFrame) else pd.DataFrame()
    intervals = hectometer_intervals.copy() if isinstance(hectometer_intervals, pd.DataFrame) else pd.DataFrame()

    short_regular = _build_short_regular_proposals(advisor)
    duplicate_names = _build_duplicate_project_names(advisor)
    object_mismatch = _build_object_family_mismatch(advisor, objects)
    knip_analysis = _build_knip_reason_analysis(advisor)
    interval_context = _build_hectometer_interval_context(intervals)
    merge_candidates = _build_merge_candidates(short_regular, duplicate_names, object_mismatch, advisor)
    summary = _build_summary_sheet(
        advisor,
        short_regular,
        duplicate_names,
        object_mismatch,
        knip_analysis,
        interval_context,
        selected_road,
        app_version,
    )

    return {
        "Samenvatting": summary,
        "Kandidaat_samenvoegen": merge_candidates,
        "Korte_reguliere_voorstellen": short_regular,
        "Dubbele_projectnamen": duplicate_names,
        "Objectfamilie_mismatch": object_mismatch,
        "Knipreden_analyse": knip_analysis,
        "Hectometerinterval_context": interval_context,
    }


def _write_dataframe_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    """Schrijf één DataFrame naar een werkblad met simpele opmaak."""
    worksheet = workbook.create_sheet(_safe_sheet_name(sheet_name))
    df = dataframe if isinstance(dataframe, pd.DataFrame) else pd.DataFrame()

    if df.empty:
        worksheet.append(["melding"])
        worksheet.append(["Geen regels voor deze categorie."])
    else:
        worksheet.append(list(df.columns))
        for _, row in df.iterrows():
            worksheet.append([row.get(column, "") for column in df.columns])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)


def build_project_calibration_workbook_bytes(
    advisor_table: pd.DataFrame | None,
    proposal_objects: pd.DataFrame | None = None,
    hectometer_intervals: pd.DataFrame | None = None,
    *,
    selected_road: str = "",
    app_version: str = "",
) -> bytes:
    """Maak het kalibratierapport als Excelbestand."""
    tables = build_project_calibration_tables(
        advisor_table,
        proposal_objects,
        hectometer_intervals,
        selected_road=selected_road,
        app_version=app_version,
    )

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name in KALIBRATIE_SHEET_ORDER:
        _write_dataframe_sheet(workbook, sheet_name, tables.get(sheet_name, pd.DataFrame()))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def project_calibration_export_filename(selected_road: str) -> str:
    """Maak de bestandsnaam voor de kalibratie-export."""
    road = sanitize_filename(selected_road or "weg")
    return f"Projectadvies_Kalibratie_{road}.xlsx"
