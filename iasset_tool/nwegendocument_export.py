"""
Concept-export in het format van het N-wegendocument.

Deze module maakt géén mutatie in het bestaande N-wegendocument. De export is
een nieuw Excelbestand met concepttabbladen per wegdeeltype. De databeheerder kan
dit als conceptwerkblad gebruiken en daarna bewust beoordelen voordat iets wordt overgenomen.

Waarom apart van Project Adviseur?
- Project Adviseur blijft de presentatielaag voor de Streamlit-app.
- Deze module vertaalt dezelfde voorstellen naar de werkvorm die databeheer
  dagelijks gebruikt: het N-wegendocument.
- De export is testbaar zonder browser en zonder iASSET-mutatie.
"""

from __future__ import annotations

import io
from collections import Counter
from collections.abc import Iterable
from typing import Any

import pandas as pd

from .utils import clean_display_value, sanitize_filename


# Interne conceptkolommen. De Excel-tabbladen gebruiken hieronder bewust
# verschillende zichtbare layouts, omdat het bestaande N-wegendocument dat ook
# doet: HRB/PW bevatten knipkolommen, FP niet.
NWEGENDOC_DATA_COLUMNS: tuple[str, ...] = (
    "onderhoudscomplex_oud",
    "statuskolom",
    "onderhoudscomplex_nieuw",
    "knip_begin",
    "knip_einde",
    "objecten",
    "locatie",
    "besteknummer",
    "documentatie",
    "verharding_begin",
    "verharding_einde",
    "verhardingsoort",
    "conservering",
    "jaar_aanleg",
    "jaar_deklaag",
    "jaar_conservering",
    "jaar_herstrating",
    "bijzonderheden",
)

HRB_VISIBLE_HEADERS: tuple[str, ...] = (
    "onderhoudscomplex oud",
    "onderhoudscomplex nieuw",
    "",
    "knip (begin)",
    "knip (einde)",
    "objecten",
    "locatie",
    "besteknummer",
    "documentatie",
    "verharding (begin)",
    "verharding (einde)",
    "verhardingsoort",
    "conservering",
    "jaar aanleg",
    "jaar deklaag",
    "jaar conservering",
    "jaar herstrating",
    "bijzonderheden",
)

HRB_VISIBLE_DATA_COLUMNS: tuple[str, ...] = (
    "onderhoudscomplex_oud",
    "statuskolom",
    "onderhoudscomplex_nieuw",
    "knip_begin",
    "knip_einde",
    "objecten",
    "locatie",
    "besteknummer",
    "documentatie",
    "verharding_begin",
    "verharding_einde",
    "verhardingsoort",
    "conservering",
    "jaar_aanleg",
    "jaar_deklaag",
    "jaar_conservering",
    "jaar_herstrating",
    "bijzonderheden",
)

PW_VISIBLE_HEADERS: tuple[str, ...] = (
    "filters obv oude complexen",
    "onderhoudscomplex oud",
    "onderhoudscomplex nieuw",
    "knip (begin)",
    "knip (einde)",
    "objecten",
    "locatie",
    "besteknummer",
    "documentatie",
    "verharding (begin)",
    "verharding (einde)",
    "verhardingsoort",
    "conservering",
    "jaar aanleg",
    "jaar deklaag",
    "jaar conservering",
    "jaar herstrating",
    "bijzonderheden",
)

PW_VISIBLE_DATA_COLUMNS: tuple[str, ...] = (
    "statuskolom",
    "onderhoudscomplex_oud",
    "onderhoudscomplex_nieuw",
    "knip_begin",
    "knip_einde",
    "objecten",
    "locatie",
    "besteknummer",
    "documentatie",
    "verharding_begin",
    "verharding_einde",
    "verhardingsoort",
    "conservering",
    "jaar_aanleg",
    "jaar_deklaag",
    "jaar_conservering",
    "jaar_herstrating",
    "bijzonderheden",
)

FP_VISIBLE_HEADERS: tuple[str, ...] = (
    "onderhoudscomplex oud",
    "onderhoudscomplex nieuw",
    "objecten",
    "locatie",
    "besteknummer",
    "verhardingsoort",
    "conservering",
    "jaar aanleg",
    "jaar deklaag",
    "jaar conservering",
    "jaar herstrating",
    "bijzonderheden",
)

FP_VISIBLE_DATA_COLUMNS: tuple[str, ...] = (
    "onderhoudscomplex_oud",
    "onderhoudscomplex_nieuw",
    "objecten",
    "locatie",
    "besteknummer",
    "verhardingsoort",
    "conservering",
    "jaar_aanleg",
    "jaar_deklaag",
    "jaar_conservering",
    "jaar_herstrating",
    "bijzonderheden",
)

NWEGENDOC_SHEET_ORDER: tuple[str, ...] = ("HRB", "PW", "FP")


def _text(value: Any) -> str:
    """Geef een veilige tekstwaarde zonder NaN/None-ruis."""
    text = clean_display_value(value).strip()
    return "" if text.lower() in {"nan", "none", "nat", "<na>"} else text


def _first_non_empty(values: Iterable[Any]) -> str:
    """Geef de eerste inhoudelijke waarde uit een reeks."""
    for value in values:
        text = _text(value)
        if text:
            return text
    return ""


def _unique_join(values: Iterable[Any], *, max_items: int = 6, separator: str = " / ") -> str:
    """Voeg unieke waarden compact samen voor een Excelcel."""
    seen: list[str] = []
    for value in values:
        text = _text(value)
        if not text or text in seen:
            continue
        seen.append(text)

    if not seen:
        return ""

    if len(seen) <= max_items:
        return separator.join(seen)

    visible = seen[:max_items]
    return separator.join(visible) + f" / … (+{len(seen) - max_items})"


def _most_common(values: Iterable[Any]) -> str:
    """Bepaal de meest voorkomende inhoudelijke waarde."""
    cleaned = [_text(value) for value in values]
    cleaned = [value for value in cleaned if value and value.lower() != "<leeg>"]
    if not cleaned:
        return ""
    return Counter(cleaned).most_common(1)[0][0]


def _number(value: Any) -> float | None:
    """Lees een getal robuust uit CSV-/DataFramewaarden."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    try:
        number = float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None
    return number if number == number else None


def _as_nwegendoc_meter_value(km_value: Any, *, fallback_meter_value: Any = None) -> int | float | str:
    """
    Schrijf knipwaarden in de schaal van het N-wegendocument.

    De projectas-engine rekent intern met relatieve route-meters op de gekozen
    wegas (`fysiek_begin_m`/`fysiek_eind_m`). Die waarden zijn nuttig voor GIS,
    maar niet vergelijkbaar met het handmatige N-wegendocument. Daar worden de
    knippen als hectometreringsmeters vastgelegd: 25.800 km wordt dus 25800.

    Daarom gebruikt deze export primair `fysiek_begin_km`/`fysiek_eind_km` en
    vermenigvuldigt die met 1000. Alleen wanneer de km-waarde ontbreekt, vallen
    we terug op de oude meterwaarde. Zo blijft de export robuust bij oudere of
    onvolledige tussenbestanden.
    """
    km_number = _number(km_value)
    if km_number is not None:
        meter_number = km_number * 1000.0
    else:
        meter_number = _number(fallback_meter_value)

    if meter_number is None:
        return ""

    rounded = round(meter_number)
    if abs(meter_number - rounded) < 0.001:
        return int(rounded)
    return round(meter_number, 3)


def _parse_technisch_profiel(profile: Any) -> dict[str, str]:
    """Parseer 'Veld=waarde, Veld=waarde' uit de projectas-engine."""
    text = _text(profile)
    parsed: dict[str, str] = {}
    if not text:
        return parsed

    for part in text.split(","):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        parsed[key.strip()] = value.strip()

    return parsed


def _value_from_objects_or_profile(
    proposal_row: pd.Series,
    object_rows: pd.DataFrame,
    object_column: str,
    profile_key: str | None = None,
) -> str:
    """Vul een N-wegendocumentwaarde vanuit objecttoewijzing met profiel-fallback."""
    if not object_rows.empty and object_column in object_rows.columns:
        value = _most_common(object_rows[object_column])
        if value:
            return value

    if profile_key:
        parsed = _parse_technisch_profiel(proposal_row.get("technisch_profiel", ""))
        value = _text(parsed.get(profile_key, ""))
        if value and value.lower() != "<leeg>":
            return value

    return ""


def _classify_sheet(row: pd.Series) -> str:
    """
    Bepaal op welk concepttabblad het voorstel hoort.

    In het N-wegendocument staan hoofdrijbanen, parallelwegen en fietspaden in
    eigen tabbladen. Busbanen en landbouwpaden worden voor deze concept-export
    bij PW geplaatst, omdat ze in het werkproces op het parallelweg-tabblad
    worden genoteerd.
    """
    project_type = _text(row.get("project_type", "")).upper()
    project_family = _text(row.get("project_family", "")).upper()
    proposed = _text(row.get("onderhoudsproject_voorgesteld", "")).upper()
    combined = " ".join(part for part in (project_type, project_family, proposed) if part)

    if "FP" in combined:
        return "FP"
    if project_type.startswith("HRB") or project_family == "HRB" or "-HRB" in proposed:
        return "HRB"

    # Parallelwegen, busbanen, landbouwpaden en overige niet-HRB/FP voorstellen
    # komen in dit concept bij het PW-tabblad.
    return "PW"


def _split_proposal_ids(value: Any) -> list[str]:
    """Lees één of meer voorstel-id's uit een samengestelde waarde."""
    text = _text(value)
    if not text:
        return []
    result: list[str] = []
    for part in text.replace(",", ";").replace("|", ";").split(";"):
        clean = part.strip()
        if clean and clean not in result:
            result.append(clean)
    return result


def _proposal_objects_for_ids(proposal_objects: pd.DataFrame, proposal_ids: Iterable[str]) -> pd.DataFrame:
    """Selecteer objecttoewijzingen voor één of meer ruwe voorstel-id's."""
    ids = {proposal_id for proposal_id in proposal_ids if proposal_id}
    if proposal_objects.empty or "voorstel_id" not in proposal_objects.columns or not ids:
        return pd.DataFrame()
    return proposal_objects[
        proposal_objects["voorstel_id"].map(_text).str.strip().isin(ids)
    ].copy()


def _proposal_objects_for_row(proposal_objects: pd.DataFrame, proposal_row: pd.Series) -> pd.DataFrame:
    """
    Selecteer objecttoewijzingen voor een conceptregel.

    Vanaf v0.37.1 kan een zichtbare onderhoudscomplexregel meerdere ruwe
    voorstel-id's bevatten in ``bron_voorstel_ids``. Voor oudere exports valt
    deze functie terug op ``voorstel_id``.
    """
    proposal_ids = _split_proposal_ids(proposal_row.get("bron_voorstel_ids", ""))
    proposal_ids.extend(_split_proposal_ids(proposal_row.get("voorstel_id", "")))

    # Verwijder duplicaten met behoud van volgorde.
    unique_ids: list[str] = []
    for proposal_id in proposal_ids:
        if proposal_id not in unique_ids:
            unique_ids.append(proposal_id)

    return _proposal_objects_for_ids(proposal_objects, unique_ids)


SPECIAL_OBJECT_MARKERS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("rotonde", ("rotonde", "turborotonde", "ovonde")),
    ("kruispunt", ("kruispunt", "kruising")),
    ("aansluiting", ("aansluiting", "oprit", "afrit")),
    ("brug", ("brug", "brugdek", "beweegbare brug")),
    ("tunnel", ("tunnel", "onderdoorgang")),
    ("viaduct", ("viaduct",)),
    ("aquaduct", ("aquaduct",)),
    ("spoorwegovergang", ("spoorwegovergang", "overweg")),
)

SPECIAL_OBJECT_COLUMNS: tuple[str, ...] = (
    "Object naam",
    "naam",
    "subthema",
    "Subthema",
    "Type onderdeel",
    "Gebruikersfunctie",
    "Thema",
    "objecttype",
    "Objecttype",
)


def _special_objects_summary(object_rows: pd.DataFrame) -> str:
    """
    Vul de kolom ``objecten`` alleen met bijzondere objecten.

    In het N-wegendocument betekent de kolom ``objecten`` niet: alle objecten
    binnen het onderhoudscomplex. De kolom is bedoeld voor herkenbare bijzondere
    objecten, zoals rotondes, kruispunten, bruggen of tunnels. Daarom schrijven
    we hier alleen iets wanneer zo'n object voorzichtig uit de paspoortvelden is
    af te leiden. Bij twijfel blijft de cel leeg.

    De volledige objecttoewijzing hoort niet in het zichtbare werkblad, maar in
    het technische objectenbestand of het tabblad ``Objecttoewijzing_data``.
    """
    if object_rows.empty:
        return ""

    found: list[str] = []
    for _, object_row in object_rows.iterrows():
        searchable_values = [
            _text(object_row.get(column, ""))
            for column in SPECIAL_OBJECT_COLUMNS
            if column in object_rows.columns
        ]
        combined = " | ".join(value.lower() for value in searchable_values if value)
        if not combined:
            continue

        for label, markers in SPECIAL_OBJECT_MARKERS:
            if not any(marker in combined for marker in markers):
                continue

            # Gebruik een herkenbare objectnaam als die zelf de marker bevat.
            # Anders gebruiken we alleen de objectcategorie, zodat we niet
            # alsnog generieke rijstrook-/fietspadnamen in deze kolom zetten.
            display_value = ""
            for name_column in ("Object naam", "naam"):
                candidate = _text(object_row.get(name_column, ""))
                if candidate and any(marker in candidate.lower() for marker in markers):
                    display_value = candidate
                    break
            if not display_value:
                display_value = label

            if display_value not in found:
                found.append(display_value)
            break

    if not found:
        return ""

    if len(found) <= 6:
        return ", ".join(found)
    return ", ".join(found[:6]) + f", … (+{len(found) - 6})"


def _build_objecttoewijzing_sheet(proposal_objects: pd.DataFrame | None) -> pd.DataFrame:
    """
    Maak een ondersteunend tabblad met objecttoewijzing.

    Het zichtbare N-wegendocument blijft compact. Deze tabel bewaart de
    objectcontext die eerder ten onrechte in de kolom ``objecten`` terechtkwam.
    Geometriekolommen worden bewust weggelaten om de Excel-export werkbaar te
    houden; de technische CSV-export blijft de volledige bron voor diepe
    diagnose.
    """
    if not isinstance(proposal_objects, pd.DataFrame) or proposal_objects.empty:
        return pd.DataFrame()

    drop_markers = ("geometry", "geom", "wkt")
    keep_columns = [
        column
        for column in proposal_objects.columns
        if not any(marker in str(column).lower() for marker in drop_markers)
    ]
    return proposal_objects[keep_columns].copy()


def _build_bijzonderheden(row: pd.Series) -> str:
    """Maak een korte, bruikbare toelichting voor de conceptregel."""
    parts: list[str] = []

    for label, column in (
        ("Werkadvies", "werkadvies"),
        ("Werklijst", "werklijst_reden"),
        ("Eindadvies", "eindadvies"),
        ("iASSET", "iassetvergelijking"),
        ("Categorie", "voorstelcategorie"),
        ("Datakwaliteit", "datakwaliteit_signalen"),
        ("Lokale afwijking", "lokale_afwijkingen"),
    ):
        value = _text(row.get(column, ""))
        if value:
            parts.append(f"{label}: {value}")

    return " | ".join(parts)


def build_nwegendocument_concept_rows(
    advisor_table: pd.DataFrame | None,
    proposal_objects: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Bouw conceptregels in dezelfde kolomvolgorde als het N-wegendocument.

    De uitvoer bevat extra hulpkolommen ``tabblad`` en ``voorstel_id`` voor de
    Excel-export. Die kolommen worden niet in de tabbladweergave zelf gezet.
    """
    if advisor_table is None or advisor_table.empty:
        columns = ["tabblad", "voorstel_id", *NWEGENDOC_DATA_COLUMNS]
        return pd.DataFrame(columns=columns)

    objects = proposal_objects if isinstance(proposal_objects, pd.DataFrame) else pd.DataFrame()
    rows: list[dict[str, Any]] = []

    for _, proposal_row in advisor_table.iterrows():
        proposal_id = _text(proposal_row.get("voorstel_id", ""))
        object_rows = _proposal_objects_for_row(objects, proposal_row)

        besteknummer = _unique_join(object_rows.get("Besteknummer", []), max_items=4) if not object_rows.empty else ""
        if not besteknummer and not object_rows.empty and "besteknummer_norm" in object_rows.columns:
            besteknummer = _unique_join(object_rows["besteknummer_norm"], max_items=4)

        verhardingsoort = _value_from_objects_or_profile(
            proposal_row,
            object_rows,
            "Soort deklaag specifiek",
            "Soort deklaag specifiek",
        )
        if not verhardingsoort:
            verhardingsoort = _value_from_objects_or_profile(
                proposal_row,
                object_rows,
                "Soort verharding_N",
                "Soort verharding_N",
            )

        row = {
            "tabblad": _classify_sheet(proposal_row),
            "voorstel_id": proposal_id,
            "onderhoudscomplex_oud": _text(proposal_row.get("bestaande_onderhoudsprojecten", "")),
            "statuskolom": "",
            "onderhoudscomplex_nieuw": _text(proposal_row.get("onderhoudsproject_voorgesteld", "")),
            "knip_begin": _as_nwegendoc_meter_value(proposal_row.get("fysiek_begin_km", ""), fallback_meter_value=proposal_row.get("fysiek_begin_m", "")),
            "knip_einde": _as_nwegendoc_meter_value(proposal_row.get("fysiek_eind_km", ""), fallback_meter_value=proposal_row.get("fysiek_eind_m", "")),
            "objecten": _special_objects_summary(object_rows),
            "locatie": "",
            "besteknummer": besteknummer,
            "documentatie": "",
            "verharding_begin": _as_nwegendoc_meter_value(proposal_row.get("fysiek_begin_km", ""), fallback_meter_value=proposal_row.get("fysiek_begin_m", "")),
            "verharding_einde": _as_nwegendoc_meter_value(proposal_row.get("fysiek_eind_km", ""), fallback_meter_value=proposal_row.get("fysiek_eind_m", "")),
            "verhardingsoort": verhardingsoort,
            "conservering": _value_from_objects_or_profile(proposal_row, object_rows, "Soort conservering", "Soort conservering"),
            "jaar_aanleg": _value_from_objects_or_profile(proposal_row, object_rows, "Jaar aanleg", "Jaar aanleg"),
            "jaar_deklaag": _value_from_objects_or_profile(proposal_row, object_rows, "Jaar deklaag", "Jaar deklaag"),
            "jaar_conservering": _value_from_objects_or_profile(proposal_row, object_rows, "Jaar conservering", "Jaar conservering"),
            "jaar_herstrating": _value_from_objects_or_profile(proposal_row, object_rows, "Jaar herstrating", "Jaar herstrating"),
            "bijzonderheden": _build_bijzonderheden(proposal_row),
        }
        rows.append(row)

    result = pd.DataFrame(rows)
    sort_columns = [column for column in ("tabblad", "knip_begin", "knip_einde", "onderhoudscomplex_nieuw") if column in result.columns]
    if sort_columns:
        result = result.sort_values(sort_columns, kind="stable").reset_index(drop=True)
    return result[["tabblad", "voorstel_id", *NWEGENDOC_DATA_COLUMNS]]


def _safe_sheet_name(name: str) -> str:
    """Maak een Excel-tabbladnaam met maximaal 31 tekens."""
    invalid = "[]:*?/\\"
    safe = "".join("_" if char in invalid else char for char in name)
    return safe[:31] or "Sheet"


def _sheet_layout(tab: str) -> tuple[tuple[str, ...], tuple[str, ...], int, str, str]:
    """
    Geef de zichtbare layout voor een N-wegendocumenttabblad.

    Retourneert:
    - headers
    - databronkolommen
    - eerste datarij
    - autofilterbereik-startcel
    - freeze_panes

    Waarom per tabblad?
    Het N-wegendocument is historisch gegroeid. HRB, PW en FP hebben niet overal
    exact dezelfde kolomindeling. De concept-export volgt die werkvorm, zodat het
    bestand naast de handmatige tabbladen te leggen is.
    """
    tab = tab.upper()
    if tab == "PW":
        return PW_VISIBLE_HEADERS, PW_VISIBLE_DATA_COLUMNS, 5, "A1", "A5"
    if tab == "FP":
        return FP_VISIBLE_HEADERS, FP_VISIBLE_DATA_COLUMNS, 2, "A1", "A2"
    return HRB_VISIBLE_HEADERS, HRB_VISIBLE_DATA_COLUMNS, 8, "A3", "A8"


def _append_status_block(worksheet, tab: str, selected_road: str) -> None:
    """Schrijf het bovenblok zoals het betreffende N-wegendocumenttabblad het verwacht."""
    tab = tab.upper()

    if tab == "FP":
        # Het fietspadtabblad in het N-wegendocument heeft een compacte layout:
        # direct de kolomkoppen, zonder statusblok en zonder knipkolommen.
        worksheet.append(list(FP_VISIBLE_HEADERS))
        return

    if tab == "PW":
        worksheet.append(list(PW_VISIBLE_HEADERS))
        worksheet.append(["", "Oud complex verwijderd uit iASSET", "Alleen paspoort gevuld"])
        worksheet.append(["", "", "Paspoort + onderhoudscomplex aangemaakt"])
        worksheet.append([])
        return

    # HRB-layout: N354 gebruikt een statusblok met oude complexen in kolom A,
    # een status/spacerkolom B en nieuwe complexen in kolom C.
    worksheet.append(["Status", "", "Concept-export Project Adviseur", selected_road])
    worksheet.append([])
    worksheet.append(list(HRB_VISIBLE_HEADERS))
    worksheet.append(["Oud onderhoudscomplex verwijderd", "", "Alleen paspoort gevuld"])
    worksheet.append(["", "", "Paspoort gevuld + nieuw onderhoudscomplex aangemaakt"])
    worksheet.append(["", "", "Maatregeltoetsdocument bijgewerkt"])
    worksheet.append([])

    # In het handmatige N354-HRB-tabblad is de kop "onderhoudscomplex nieuw"
    # visueel over kolom B en C getrokken. Dat nemen we over; de eigenlijke
    # projectnamen blijven in kolom C.
    try:
        worksheet.merge_cells("B3:C3")
    except ValueError:
        # Robuust houden bij schrijf-/templatevarianten.
        pass


def _write_nwegendoc_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    rows: pd.DataFrame,
    selected_road: str,
    *,
    tab: str,
) -> None:
    """Schrijf één concepttabblad in N-wegendocument-layout."""
    workbook = writer.book
    worksheet = workbook.create_sheet(_safe_sheet_name(sheet_name))
    writer.sheets[worksheet.title] = worksheet

    headers, data_columns, first_data_row, filter_start_cell, freeze_panes = _sheet_layout(tab)
    _append_status_block(worksheet, tab, selected_road)

    display_rows = rows[list(data_columns)] if not rows.empty else pd.DataFrame(columns=data_columns)
    for values in display_rows.itertuples(index=False, name=None):
        worksheet.append(list(values))

    # Opmaak blijft bewust eenvoudig maar herkenbaar: kopregels, bevroren rijen,
    # filter en goed leesbare kolombreedtes.
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    status_fill = PatternFill("solid", fgColor="F2F2F2")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = worksheet.max_row
    max_col = len(headers)

    if tab.upper() == "HRB":
        header_rows = (1, 3, 4, 5, 6)
    elif tab.upper() == "PW":
        header_rows = (1, 2, 3)
    else:
        header_rows = (1,)

    for row_idx in header_rows:
        if row_idx > worksheet.max_row:
            continue
        for cell in worksheet[row_idx]:
            cell.fill = header_fill if row_idx in {1, 3} else status_fill
            cell.font = Font(bold=row_idx in {1, 3})
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in worksheet.iter_rows(min_row=first_data_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if max_row >= first_data_row:
        worksheet.auto_filter.ref = f"{filter_start_cell}:{worksheet.cell(max_row, max_col).coordinate}"

    worksheet.freeze_panes = freeze_panes

    if tab.upper() == "FP":
        widths = {
            "A": 24,
            "B": 24,
            "C": 34,
            "D": 22,
            "E": 20,
            "F": 24,
            "G": 18,
            "H": 12,
            "I": 12,
            "J": 16,
            "K": 16,
            "L": 60,
        }
        note_column = "L"
    elif tab.upper() == "PW":
        widths = {
            "A": 24,
            "B": 24,
            "C": 24,
            "D": 12,
            "E": 12,
            "F": 34,
            "G": 22,
            "H": 20,
            "I": 28,
            "J": 13,
            "K": 13,
            "L": 24,
            "M": 18,
            "N": 12,
            "O": 12,
            "P": 16,
            "Q": 16,
            "R": 60,
        }
        note_column = "R"
    else:
        widths = {
            "A": 24,
            "B": 4,
            "C": 24,
            "D": 12,
            "E": 12,
            "F": 34,
            "G": 22,
            "H": 20,
            "I": 28,
            "J": 13,
            "K": 13,
            "L": 24,
            "M": 18,
            "N": 12,
            "O": 12,
            "P": 16,
            "Q": 16,
            "R": 60,
        }
        note_column = "R"

    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    if max_row >= first_data_row:
        for cell_tuple in worksheet[f"{note_column}{first_data_row}:{note_column}{max_row}"]:
            cell_tuple[0].fill = note_fill


def _write_dataframe_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """Schrijf een ondersteunend tabblad met simpele tabelopmaak."""
    safe_name = _safe_sheet_name(sheet_name)
    df_to_write = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    df_to_write.to_excel(writer, sheet_name=safe_name, index=False)

    worksheet = writer.sheets[safe_name]
    if worksheet.max_row > 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(_text(cell.value)) for cell in column_cells[:100])
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)


def build_nwegendocument_concept_workbook_bytes(
    advisor_table: pd.DataFrame | None,
    proposal_objects: pd.DataFrame | None = None,
    run_report: pd.DataFrame | None = None,
    *,
    selected_road: str = "",
    app_version: str = "",
) -> bytes:
    """
    Maak een Excelbestand met concepttabbladen in N-wegendocument-format.

    Dit bestand is bedoeld om naast het bestaande N-wegendocument te leggen. Het
    overschrijft geen bronbestand en is geen directe iASSET-mutatie.

    Vanaf v0.37.1 mag ``advisor_table`` ook de zichtbare
    onderhoudscomplexlaag zijn. Als daarin ``in_concept_nwegendocument`` staat,
    worden alleen die zichtbare regels in HRB/PW/FP gezet. De uitgesloten regels
    blijven als controlepunten beschikbaar in ``Controlepunten_data``.
    """
    source_table = advisor_table if isinstance(advisor_table, pd.DataFrame) else pd.DataFrame()
    control_points = pd.DataFrame()
    if not source_table.empty and "in_concept_nwegendocument" in source_table.columns:
        include_mask = source_table["in_concept_nwegendocument"].map(
            lambda value: str(value).strip().lower() in {"true", "waar", "ja", "yes", "1"}
            if not isinstance(value, bool)
            else value
        )
        control_points = source_table[~include_mask].copy()
        source_table = source_table[include_mask].copy()

    concept_rows = build_nwegendocument_concept_rows(source_table, proposal_objects)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Bewaar het standaardblad tijdelijk. We verwijderen het pas nadat
        # pandas/openpyxl minimaal één echt werkblad heeft aangemaakt.
        default_sheet = writer.book.active

        summary_rows = [
            {"onderdeel": "Doel", "waarde": "Concepttabblad in N-wegendocument-format"},
            {"onderdeel": "Weg", "waarde": selected_road or "onbekend"},
            {"onderdeel": "App-versie", "waarde": app_version or "onbekend"},
            {
                "onderdeel": "Belangrijk",
                "waarde": "Concept-export: controleer altijd met kaartbeeld, actuele iASSET-data en beschikbare broninformatie voordat je gegevens verwerkt. Het concept-N-wegendocument is een werkblad, geen waarheid.",
            },
            {
                "onderdeel": "Niet automatisch gevuld",
                "waarde": "locatie, documentatie en menselijke opmerkingen blijven grotendeels handmatig.",
            },
            {
                "onderdeel": "Objecten-kolom",
                "waarde": "Alleen bijzondere objecten worden zichtbaar gevuld; volledige objecttoewijzing staat apart.",
            },
        ]
        _write_dataframe_sheet(writer, "Samenvatting", pd.DataFrame(summary_rows))

        if (
            default_sheet is not None
            and default_sheet.title in writer.book.sheetnames
            and default_sheet.max_row == 1
            and default_sheet.max_column == 1
            and default_sheet["A1"].value is None
            and len(writer.book.worksheets) > 1
        ):
            writer.book.remove(default_sheet)

        for tab in NWEGENDOC_SHEET_ORDER:
            rows_for_tab = concept_rows[concept_rows["tabblad"] == tab].copy() if not concept_rows.empty else concept_rows.copy()
            if rows_for_tab.empty and tab != "HRB":
                # Maak PW/FP alleen aan als ze echt voorkomen. HRB blijft bestaan
                # zodat een lege run voor een hoofdrijbaan herkenbaar is.
                continue
            _write_nwegendoc_sheet(writer, f"{selected_road or 'Weg'} ({tab})", rows_for_tab, selected_road, tab=tab)

        not_auto = pd.DataFrame(
            [
                {
                    "veld": "locatie",
                    "toelichting": "Herkenbare punten zoals brug, kruising of rotonde worden niet betrouwbaar uit de paspoortdata afgeleid.",
                },
                {
                    "veld": "documentatie",
                    "toelichting": "Dielplak-/besteklinks zitten niet altijd in de iASSET-export en blijven daarom leeg tenzij brondata dit bevat.",
                },
                {
                    "veld": "objecten",
                    "toelichting": "Alleen bijzondere objecten zoals rotonde, kruispunt, brug, tunnel of viaduct. Geen volledige objectlijst.",
                },
                {
                    "veld": "micro-/eindzone",
                    "toelichting": "Deze regels staan in bijzonderheden en zijn niet bedoeld als regulier onderhoudscomplex zonder beoordeling.",
                },
                {
                    "veld": "buiten ijkbereik",
                    "toelichting": "Niet overnemen zonder handmatige kaart- en broncontrole.",
                },
                {
                    "veld": "FP-tabblad",
                    "toelichting": "Het fietspadtabblad volgt het compacte N-wegendocument-format zonder knipkolommen.",
                },
            ]
        )
        _write_dataframe_sheet(writer, "Niet automatisch gevuld", not_auto)

        if isinstance(run_report, pd.DataFrame) and not run_report.empty:
            _write_dataframe_sheet(writer, "Runrapport", run_report)

        objecttoewijzing = _build_objecttoewijzing_sheet(proposal_objects)
        if not objecttoewijzing.empty:
            _write_dataframe_sheet(writer, "Objecttoewijzing_data", objecttoewijzing)

        if not control_points.empty:
            _write_dataframe_sheet(writer, "Controlepunten_data", control_points)

        if isinstance(advisor_table, pd.DataFrame) and not advisor_table.empty:
            zichtbare_columns = [
                column
                for column in (
                    "zichtbaar_complex_id",
                    "zichtbare_status",
                    "zichtbare_klasse",
                    "in_concept_nwegendocument",
                    "onderhoudsproject_voorgesteld",
                    "project_type",
                    "fysiek_begin_km",
                    "fysiek_eind_km",
                    "fysiek_lengte_m",
                    "aantal_ruwe_voorstellen",
                    "bron_voorstel_ids",
                    "objectfamilie_mismatch_aantal",
                    "bijzonderheden",
                )
                if column in advisor_table.columns
            ]
            if zichtbare_columns:
                _write_dataframe_sheet(writer, "Zichtbare_complexen_data", advisor_table[zichtbare_columns].copy())

        if not concept_rows.empty:
            _write_dataframe_sheet(writer, "Conceptregels_data", concept_rows)

    output.seek(0)
    return output.getvalue()


def nwegendocument_export_filename(selected_road: str) -> str:
    """Maak een consistente bestandsnaam voor de concept-export."""
    road = sanitize_filename(selected_road or "weg")
    return f"Projectadvies_Nwegendocument_{road}.xlsx"
